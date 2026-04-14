"""
stage8_hr_payroll.py — Import HR employees, contracts, payroll runs, and payroll entries.

Sources:
  - OldData/7-April-2026/PAYROLL Mar -2 2026.xlsx  (sheet "Countrywise") — 141 WPS employees
  - OldData/7-April-2026/PAYROLL Mar 2026.xlsx     (sheet "Countrywise") — 1 WPS employee
  - OldData/7-April-2026/Salary - Mar 2026 Withou Visa Cash.xlsx — 5 farm cash sheets

New / extended collections:
  - employees          (extend existing — 1 doc already present)
  - employee_contracts (new collection)
  - payroll_runs       (new collection)
  - payroll_entries    (new collection)

Idempotency: upsert keyed on metadata.legacyRef for all collections.

Important notes on Employee model compatibility:
  - EmployeeBase requires: firstName, lastName, email, department, position,
    hireDate, status (all required, no defaults for first four + hireDate).
  - The HR model does NOT have paymentType / iban / passportNumber / farmId
    fields at the Pydantic level. Extra fields are stored in metadata for
    traceability but the core doc satisfies EmployeeBase validation.
  - Contract model uses salary: float > 0; for 0-salary VISA workers we set
    salary to 0.01 with a flag, or store contractual=0 in metadata and skip
    Pydantic validation on the contract doc (separate collection shape).
  - We do NOT run Pydantic validation on payroll_runs / payroll_entries since
    those are new collections with no existing Pydantic model.

Run:  python stage8_hr_payroll.py [--dry-run] [--reset]
"""

from __future__ import annotations

import re
import sys
import uuid
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from common import (
    DIVISION_ID,
    MIGRATION_TAG,
    ORGANIZATION_ID,
    deterministic_uuid,
    get_db,
    make_arg_parser,
    make_logger,
    print_summary,
    reset_migration_data,
    upsert_by_legacy_ref,
    utcnow,
)

STAGE = "stage8_hr_payroll"
RESET_COLLECTIONS = [
    "employees",
    "employee_contracts",
    "payroll_runs",
    "payroll_entries",
]

ADMIN_UUID: str = "bff26b8f-5ce9-49b2-9126-86174eaea823"

# Path constants
DATA_DIR: Path = Path(__file__).parent.parent.parent.parent / "OldData" / "7-April-2026"
WPS1_FILE: Path = DATA_DIR / "PAYROLL Mar -2 2026.xlsx"
WPS2_FILE: Path = DATA_DIR / "PAYROLL Mar 2026.xlsx"
CASH_FILE: Path = DATA_DIR / "Salary - Mar 2026 Withou Visa Cash.xlsx"

WPS_COMPANY: str = "OFFICE OF SHAIKH"

# March 2026 payroll run date (date signed off per "All PV" sheet label)
PAYROLL_RUN_DATE: str = "2026-04-02"
PAYROLL_PERIOD_YEAR: int = 2026
PAYROLL_PERIOD_MONTH: int = 3
PAYROLL_PERIOD_LABEL: str = "March 2026"

# Farm name → UUID map (from migrated farms)
# Silal Upgrade maps to "Silal Upgrade Farm" in the farms collection
FARM_NAME_MAP: dict[str, str] = {
    "al ain": "042ab6a6-74c2-58ca-84ec-73dd3186b0d9",
    "liwa": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
    "al khazana": "2b34823c-ac2d-58c8-a57d-ed2c8fbbdf81",
    "al khazna": "2b34823c-ac2d-58c8-a57d-ed2c8fbbdf81",
    "khaza": "2b34823c-ac2d-58c8-a57d-ed2c8fbbdf81",
    "al wagen": "b2c23ace-a0db-549c-9872-28f29658a2f3",
    "al wagan": "b2c23ace-a0db-549c-9872-28f29658a2f3",
    "silal upgrade": "651103f1-1967-5261-9a38-373411f4fdfa",
    "silal upgrade farm": "651103f1-1967-5261-9a38-373411f4fdfa",
}

# Nationality code → full string
NATIONALITY_CODE_MAP: dict[str, str] = {
    "BD": "Bangladeshi",
    "SD": "Sudanese",
    "LB": "Lebanese",
    "IN": "Indian",
    "PK": "Pakistani",
    "EG": "Egyptian",
    "ET": "Ethiopian",
    "NG": "Nigerian",
    "PH": "Filipino",
    "NP": "Nepali",
    "LK": "Sri Lankan",
    "KE": "Kenyan",
    "TZ": "Tanzanian",
    "UG": "Ugandan",
    "GH": "Ghanaian",
    "AE": "Emirati",
    "JO": "Jordanian",
    "SY": "Syrian",
    "IQ": "Iraqi",
    "YE": "Yemeni",
    "MA": "Moroccan",
    "TN": "Tunisian",
    "DZ": "Algerian",
    "SO": "Somali",
    "ER": "Eritrean",
}


# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------


def split_name(full_name: str) -> tuple[str, str]:
    """
    Split a full name into (firstName, lastName).

    Uses first word as firstName and remainder as lastName.
    If only one word, lastName = firstName.

    Args:
        full_name: raw full name string

    Returns:
        (firstName, lastName) tuple
    """
    parts = full_name.strip().split()
    if not parts:
        return ("Unknown", "Unknown")
    if len(parts) == 1:
        return (parts[0], parts[0])
    return (parts[0], " ".join(parts[1:]))


def normalize_name(name: str) -> str:
    """
    Return lowercase stripped name for dedup key.

    Args:
        name: raw name

    Returns:
        normalized string
    """
    return re.sub(r"\s+", " ", name.strip().lower())


def normalize_eid(eid: Any) -> Optional[str]:
    """
    Normalize Emirates ID to digit-only string.

    Args:
        eid: raw EID value (string or None)

    Returns:
        digit string or None if invalid / placeholder
    """
    if not eid:
        return None
    s = str(eid).strip().replace("-", "").replace(" ", "")
    if not s.isdigit():
        return None
    if len(s) < 10:
        # Reason: "No E ID" and "0" are placeholders, not real EIDs
        return None
    return s


def normalize_nationality(raw: Any, code: Any = None) -> Optional[str]:
    """
    Resolve nationality to a clean string.

    Args:
        raw: nationality string from data (may be None)
        code: 2-letter code from WPS payroll Code column

    Returns:
        normalized nationality string or None
    """
    if code and str(code).strip().upper() in NATIONALITY_CODE_MAP:
        return NATIONALITY_CODE_MAP[str(code).strip().upper()]
    if raw:
        s = str(raw).strip().title()
        if s.lower() in ("bangladeshi", "bangladesh"):
            return "Bangladeshi"
        if s.lower() in ("sudanese", "sudan"):
            return "Sudanese"
        if s.lower() in ("ethiopian", "ethiopia"):
            return "Ethiopian"
        if s.lower() in ("egyptian", "egypt"):
            return "Egyptian"
        if s.lower() in ("lebanese", "lebanon"):
            return "Lebanese"
        return s
    return None


def resolve_farm(location_hint: Optional[str]) -> Optional[str]:
    """
    Map a location string to a farmId UUID.

    Args:
        location_hint: raw location / sheet name

    Returns:
        farmId UUID string or None
    """
    if not location_hint:
        return None
    key = normalize_name(location_hint).replace(" farm", "").strip()
    return FARM_NAME_MAP.get(key) or FARM_NAME_MAP.get(location_hint.strip().lower())


def parse_date(val: Any) -> Optional[date]:
    """
    Parse a date value from xlsx (datetime object, string, or None).

    Args:
        val: raw cell value

    Returns:
        date object or None
    """
    # Returns datetime (not date) so pymongo can encode it
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def make_employee_email(name: str, emp_uuid: str) -> str:
    """
    Generate a synthetic email for migration-imported employees.

    Args:
        name: employee full name
        emp_uuid: deterministic UUID

    Returns:
        synthetic email string
    """
    slug = re.sub(r"[^a-z0-9]", ".", normalize_name(name))
    slug = re.sub(r"\.{2,}", ".", slug).strip(".")
    # Truncate FIRST, then re-collapse/trim — otherwise slug[:30] can end with '.'
    # and the f-string concatenation below produces '..' (an invalid email per RFC 5322).
    slug = re.sub(r"\.{2,}", ".", slug[:30]).strip(".")
    short_id = emp_uuid.split("-")[0]
    return f"emp.{slug}.{short_id}@a64farms.ae"


# ---------------------------------------------------------------------------
# Row-level builders
# ---------------------------------------------------------------------------


def legacy_ref_for_employee(name: str, eid: Optional[str]) -> str:
    """
    Build a stable legacy reference key for an employee.

    Prefers EID when present; falls back to normalized name.

    Args:
        name: full name (normalized)
        eid: normalized emirates ID or None

    Returns:
        stable string key
    """
    if eid:
        return f"eid:{eid}"
    return f"name:{normalize_name(name)}"


def build_employee_doc(
    full_name: str,
    eid: Optional[str],
    passport: Optional[str],
    visa_no: Optional[str],
    bank_account: Optional[str],
    dob: Optional[date],
    nationality: Optional[str],
    address: Optional[str],
    join_date: Optional[date],
    designation: Optional[str],
    department: Optional[str],
    farm_id: Optional[str],
    payment_type: str,
    payment_company: Optional[str],
    status_raw: Optional[str],
    source_sheet: str,
) -> dict[str, Any]:
    """
    Build an employees collection document.

    Args:
        full_name: employee full name
        eid: normalized Emirates ID
        passport: passport number
        visa_no: visa number
        bank_account: IBAN / account number
        dob: date of birth
        nationality: nationality string
        address: address
        join_date: hire/join date
        designation: job title/designation
        department: department
        farm_id: assigned farm UUID
        payment_type: "wps" | "cash"
        payment_company: paying company name
        status_raw: raw STATUS column value (from Silal Upgrade)
        source_sheet: original sheet name for traceability

    Returns:
        employees document dict
    """
    now = utcnow()
    legacy_ref = legacy_ref_for_employee(full_name, eid)
    emp_uuid = deterministic_uuid("employee", legacy_ref)
    first, last = split_name(full_name)
    email = make_employee_email(full_name, emp_uuid)

    # Hire date: fall back to a migration-era sentinel if not available
    hire_date_val = join_date or datetime(2026, 3, 1)

    # Department: use provided value, else default based on payment type
    dept = (department or "Farm Operations").strip()[:100]

    # Position/designation
    position = (designation or "Labor").strip()[:100]

    # Status from STATUS column — "VISA" and "NO" are both active employees
    status = "active"

    return {
        "employeeId": emp_uuid,
        "employeeCode": None,  # Not assigned in migration batch
        "firstName": first[:100],
        "lastName": last[:100],
        "arabicFirstName": None,
        "arabicMiddleName": None,
        "arabicLastName": None,
        "email": email,
        "phone": None,
        "department": dept,
        "position": position,
        "hireDate": hire_date_val,
        "status": status,
        "gender": None,
        "nationality": nationality,
        "maritalStatus": None,
        "emiratesId": eid,
        "visaIssuancePlace": None,
        "emergencyContact": None,
        # Extended migration fields stored as flat keys on the doc
        "fullName": full_name.strip(),
        "passportNumber": passport,
        "visaNumber": visa_no,
        "iban": bank_account,
        "dateOfBirth": dob,
        "address": address,
        "joinDate": join_date,
        "farmId": farm_id,
        "paymentType": payment_type,
        "paymentCompany": payment_company,
        "isActive": True,
        "createdBy": ADMIN_UUID,
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": legacy_ref,
            "sourceSheet": source_sheet,
            "rawStatusColumn": status_raw,
            "paymentType": payment_type,
        },
    }


def build_contract_doc(
    employee_id: str,
    monthly_salary: float,
    payment_type: str,
    payment_company: Optional[str],
    farm_id: Optional[str],
    legacy_ref: str,
) -> dict[str, Any]:
    """
    Build an employee_contracts collection document.

    Note: We do NOT validate against the existing Contract Pydantic model because:
    - Contract.salary has `gt=0` constraint; VISA workers legitimately have 0 salary.
    - Contract.type uses ContractType enum (full_time | part_time | contractor | intern)
      which does not align with "wps" | "cash" payment classification.
    We store payroll-specific fields directly and use full_time as the closest proxy.

    Args:
        employee_id: employee UUID
        monthly_salary: gross monthly salary in AED
        payment_type: "wps" | "cash"
        payment_company: paying company name
        farm_id: assigned farm UUID
        legacy_ref: employee legacy ref for deterministic UUID

    Returns:
        employee_contracts document dict
    """
    now = utcnow()
    contract_ref = f"contract:{legacy_ref}"
    contract_uuid = deterministic_uuid("employee_contract", contract_ref)

    # Pydantic Contract.salary requires > 0; we store contractMonthlySalary
    # as a migration-specific field to handle 0-salary VISA workers.
    return {
        "contractId": contract_uuid,
        "employeeId": employee_id,
        # Map to closest ContractType enum value
        "type": "full_time",
        "contractType": payment_type,  # payroll classification
        "startDate": datetime(2026, 3, 1),
        "endDate": None,
        # Reason: Pydantic salary gt=0; use 0.01 as minimum for VISA workers
        # contractMonthlySalary holds the real value (may be 0)
        "salary": max(float(monthly_salary), 0.01),
        "contractMonthlySalary": float(monthly_salary),
        "currency": "AED",
        "benefits": [],
        "status": "active",
        "documentUrl": None,
        "employmentStatus": "active",
        "paymentCompany": payment_company,
        "farmId": farm_id,
        "createdBy": ADMIN_UUID,
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": contract_ref,
            "paymentType": payment_type,
            "effectiveFrom": "2026-03-01",
        },
    }


def build_payroll_run_doc(
    run_id: str,
    total_amount: float,
    total_employees: int,
    by_farm: dict,
    by_company: dict,
    by_payment_type: dict,
) -> dict[str, Any]:
    """
    Build a payroll_runs collection document for March 2026.

    Args:
        run_id: deterministic UUID for the run
        total_amount: total payroll amount in AED
        total_employees: employee count
        by_farm: {farmId: {count, amount}} breakdown
        by_company: {companyName: {count, amount}} breakdown
        by_payment_type: {wps: {count, amount}, cash: {count, amount}}

    Returns:
        payroll_runs document dict
    """
    now = utcnow()
    return {
        "runId": run_id,
        "runDate": datetime(2026, 4, 2),
        "period": {
            "year": PAYROLL_PERIOD_YEAR,
            "month": PAYROLL_PERIOD_MONTH,
            "label": PAYROLL_PERIOD_LABEL,
        },
        "totalAmount": round(total_amount, 2),
        "totalEmployees": total_employees,
        "byFarm": by_farm,
        "byCompany": by_company,
        "byPaymentType": by_payment_type,
        "status": "completed",
        "createdBy": ADMIN_UUID,
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": f"payroll_run:march_2026",
            "sourceSheets": [
                "PAYROLL Mar -2 2026 / Countrywise",
                "PAYROLL Mar 2026 / Countrywise",
                "Salary - Mar 2026 Withou Visa Cash / Al Ain -Salary",
                "Salary - Mar 2026 Withou Visa Cash / Liwa",
                "Salary - Mar 2026 Withou Visa Cash / khaza",
                "Salary - Mar 2026 Withou Visa Cash / Al Wagen",
                "Salary - Mar 2026 Withou Visa Cash / Silal Upgrade",
            ],
        },
    }


def build_payroll_entry_doc(
    run_id: str,
    employee_id: str,
    employee_name: str,
    gross_salary: float,
    deductions: float,
    net_salary: float,
    farm_id: Optional[str],
    farm_name: Optional[str],
    payment_type: str,
    payment_company: Optional[str],
    status: str,
    source_sheet: str,
) -> dict[str, Any]:
    """
    Build a payroll_entries collection document.

    Args:
        run_id: payroll run UUID
        employee_id: employee UUID
        employee_name: denormalized full name
        gross_salary: gross pay AED
        deductions: deductions AED
        net_salary: net pay AED
        farm_id: farm UUID or None
        farm_name: farm name string
        payment_type: "wps" | "cash"
        payment_company: paying company
        status: entry status string
        source_sheet: original sheet name

    Returns:
        payroll_entries document dict
    """
    now = utcnow()
    legacy_ref = f"payroll_entry:march_2026:{employee_id}"
    entry_uuid = deterministic_uuid("payroll_entry", legacy_ref)

    return {
        "entryId": entry_uuid,
        "runId": run_id,
        "employeeId": employee_id,
        "employeeName": employee_name,
        "grossSalary": round(float(gross_salary), 2),
        "deductions": round(float(deductions or 0), 2),
        "netSalary": round(float(net_salary or gross_salary), 2),
        "totalSalary": round(float(net_salary or gross_salary), 2),
        "farmId": farm_id,
        "farmName": farm_name,
        "paymentType": payment_type,
        "paymentCompany": payment_company,
        "status": status,
        "createdBy": ADMIN_UUID,
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": legacy_ref,
            "sourceSheet": source_sheet,
        },
    }


# ---------------------------------------------------------------------------
# Excel loaders
# ---------------------------------------------------------------------------


def load_wps_sheet(filepath: Path, source_label: str) -> list[dict[str, Any]]:
    """
    Load a WPS payroll Countrywise sheet.

    Headers at row 6; data from row 7.
    Stops at first row where Sr No is non-integer.

    Args:
        filepath: path to xlsx file
        source_label: label for source_sheet field

    Returns:
        list of parsed employee row dicts
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Countrywise"]
    rows: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=7, values_only=True):
        sr = row[0]
        if sr is None or not isinstance(sr, (int, float)):
            continue
        eid_raw = row[1]
        name = str(row[2]).strip() if row[2] else ""
        if not name:
            continue
        passport = str(row[3]).strip() if row[3] else None
        visa_no = str(row[4]).strip() if row[4] else None
        account = str(row[5]).strip() if row[5] else None
        salary = float(row[6]) if row[6] is not None else 0.0
        address = str(row[7]).strip() if row[7] else None
        dob = parse_date(row[8])
        nationality_raw = str(row[9]).strip() if row[9] else None
        code = str(row[10]).strip() if row[10] else None

        eid = normalize_eid(eid_raw)
        nationality = normalize_nationality(nationality_raw, code)

        rows.append(
            {
                "full_name": name,
                "eid": eid,
                "passport": passport,
                "visa_no": visa_no,
                "bank_account": account if account and account != "0" else None,
                "salary": salary,
                "address": address,
                "dob": dob,
                "nationality": nationality,
                "join_date": None,
                "designation": None,
                "department": None,
                "farm_id": None,  # WPS = company-wide, no specific farm
                "payment_type": "wps",
                "payment_company": WPS_COMPANY,
                "status_raw": None,
                "source_sheet": source_label,
                "deductions": 0.0,
                "net_salary": salary,
            }
        )
    return rows


def load_cash_sheet_al_ain() -> list[dict[str, Any]]:
    """Load Al Ain -Salary sheet (no header row, headers at row 3, data from row 4)."""
    wb = openpyxl.load_workbook(CASH_FILE, data_only=True)
    ws = wb["Al Ain -Salary"]
    # Headers: SL NO., EMP Id, Employee Name, IBAN No, Gross, Deduction, Total Salary, Status
    rows = []
    farm_id = FARM_NAME_MAP["al ain"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        sl = row[0]
        if sl is None or not isinstance(sl, (int, float)):
            continue
        name = str(row[2]).strip() if row[2] else ""
        if not name:
            continue
        iban_raw = row[3]
        gross = float(row[4]) if row[4] is not None else 0.0
        deduction = float(row[5]) if row[5] is not None else 0.0
        total = float(row[6]) if row[6] is not None else gross
        status_raw = str(row[7]).strip() if row[7] else None
        rows.append(
            {
                "full_name": name,
                "eid": None,
                "passport": None,
                "visa_no": None,
                "bank_account": str(iban_raw) if iban_raw and str(iban_raw) != "0" else None,
                "salary": gross,
                "address": None,
                "dob": None,
                "nationality": None,
                "join_date": None,
                "designation": None,
                "department": None,
                "farm_id": farm_id,
                "payment_type": "cash",
                "payment_company": None,
                "status_raw": status_raw,
                "source_sheet": "Al Ain -Salary",
                "deductions": deduction,
                "net_salary": total,
            }
        )
    return rows


def load_cash_sheet_liwa() -> list[dict[str, Any]]:
    """Load Liwa sheet (headers at row 2, data from row 3)."""
    wb = openpyxl.load_workbook(CASH_FILE, data_only=True)
    ws = wb["Liwa"]
    # Headers: Sr No, EMP Id, Employee Name, IBAN No, VAC, Days, Gross, Deduction, Total Salary, Status
    rows = []
    farm_id = FARM_NAME_MAP["liwa"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        sl = row[0]
        if sl is None or not isinstance(sl, (int, float)):
            continue
        name = str(row[2]).strip() if row[2] else ""
        if not name:
            continue
        iban_raw = row[3]
        gross = float(row[6]) if row[6] is not None else 0.0
        deduction = float(row[7]) if row[7] is not None else 0.0
        total = float(row[8]) if row[8] is not None else gross
        status_raw = str(row[9]).strip() if row[9] else None
        rows.append(
            {
                "full_name": name,
                "eid": None,
                "passport": None,
                "visa_no": None,
                "bank_account": str(iban_raw) if iban_raw and str(iban_raw) != "0" else None,
                "salary": gross,
                "address": None,
                "dob": None,
                "nationality": None,
                "join_date": None,
                "designation": None,
                "department": None,
                "farm_id": farm_id,
                "payment_type": "cash",
                "payment_company": None,
                "status_raw": status_raw,
                "source_sheet": "Liwa",
                "deductions": deduction,
                "net_salary": total,
            }
        )
    return rows


def load_cash_sheet_khaza() -> list[dict[str, Any]]:
    """Load khaza (Al Khazana) sheet (headers at row 2, data from row 3)."""
    wb = openpyxl.load_workbook(CASH_FILE, data_only=True)
    ws = wb["khaza"]
    # Headers: SL NO., EMP Id, Employee Name, Dept, IBAN No, Gross Salary, Deduction, Total Salary, Status
    rows = []
    farm_id = FARM_NAME_MAP["al khazana"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        sl = row[0]
        if sl is None or not isinstance(sl, (int, float)):
            continue
        name = str(row[2]).strip() if row[2] else ""
        if not name:
            continue
        dept = str(row[3]).strip() if row[3] else None
        iban_raw = row[4]
        gross = float(row[5]) if row[5] is not None else 0.0
        deduction = float(row[6]) if row[6] is not None else 0.0
        total = float(row[7]) if row[7] is not None else gross
        status_raw = str(row[8]).strip() if row[8] else None
        rows.append(
            {
                "full_name": name,
                "eid": None,
                "passport": None,
                "visa_no": None,
                "bank_account": str(iban_raw) if iban_raw and str(iban_raw) != "0" else None,
                "salary": gross,
                "address": None,
                "dob": None,
                "nationality": None,
                "join_date": None,
                "designation": None,
                "department": dept,
                "farm_id": farm_id,
                "payment_type": "cash",
                "payment_company": None,
                "status_raw": status_raw,
                "source_sheet": "khaza",
                "deductions": deduction,
                "net_salary": total,
            }
        )
    return rows


def load_cash_sheet_al_wagen() -> list[dict[str, Any]]:
    """Load Al Wagen sheet (headers at row 2, data from row 3)."""
    wb = openpyxl.load_workbook(CASH_FILE, data_only=True)
    ws = wb["Al Wagen"]
    # Headers: Sr No., EMP Id, Employee Name, IBAN No, Working Days, Total Salary, Status
    rows = []
    farm_id = FARM_NAME_MAP["al wagen"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        sl = row[0]
        if sl is None or not isinstance(sl, (int, float)):
            continue
        name = str(row[2]).strip() if row[2] else ""
        if not name:
            continue
        iban_raw = row[3]
        total = float(row[5]) if row[5] is not None else 0.0
        status_raw = str(row[6]).strip() if row[6] else None
        rows.append(
            {
                "full_name": name,
                "eid": None,
                "passport": None,
                "visa_no": None,
                "bank_account": str(iban_raw) if iban_raw and str(iban_raw) != "0" else None,
                "salary": total,
                "address": None,
                "dob": None,
                "nationality": None,
                "join_date": None,
                "designation": None,
                "department": None,
                "farm_id": farm_id,
                "payment_type": "cash",
                "payment_company": None,
                "status_raw": status_raw,
                "source_sheet": "Al Wagen",
                "deductions": 0.0,
                "net_salary": total,
            }
        )
    return rows


def load_cash_sheet_silal_upgrade() -> list[dict[str, Any]]:
    """
    Load Silal Upgrade sheet (headers at row 2, data from row 3).

    This sheet has the richest data: Location, Farm No, Doc No,
    Employee Name, STATUS, Nationality, Date of Birth, Designation,
    Join Date, Days, Salary, Deduction, Total Salary.

    VISA workers with Salary=0 are included as active employees.
    """
    wb = openpyxl.load_workbook(CASH_FILE, data_only=True)
    ws = wb["Silal Upgrade"]
    # Headers row 2: SL NO., Location, Farm No, Doc No, Employee Name, STATUS,
    #   Nationality, Date of Birth, Designation, Join Date, Days, Salary,
    #   Deduction- Fines Visa / Insurance, Total Salary
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        sl = row[0]
        if sl is None or not isinstance(sl, (int, float)):
            continue
        location = str(row[1]).strip() if row[1] else None
        name = str(row[4]).strip() if row[4] else ""
        if not name or name.startswith("…"):
            continue
        status_raw = str(row[5]).strip() if row[5] else None
        nationality_raw = str(row[6]).strip() if row[6] else None
        dob = parse_date(row[7])
        designation = str(row[8]).strip() if row[8] else "Labor"
        join_date = parse_date(row[9])
        salary = float(row[11]) if row[11] is not None else 0.0
        deduction = float(row[12]) if row[12] is not None else 0.0
        total = float(row[13]) if row[13] is not None else salary

        farm_id = resolve_farm(location) if location else FARM_NAME_MAP.get("silal upgrade")
        nationality = normalize_nationality(nationality_raw)

        rows.append(
            {
                "full_name": name,
                "eid": None,
                "passport": None,
                "visa_no": None,
                "bank_account": None,
                "salary": salary,
                "address": None,
                "dob": dob,
                "nationality": nationality,
                "join_date": join_date,
                "designation": designation,
                "department": None,
                "farm_id": farm_id,
                "payment_type": "cash",
                "payment_company": None,
                "status_raw": status_raw,
                "source_sheet": "Silal Upgrade",
                "deductions": deduction,
                "net_salary": total,
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Deduplication logic
# ---------------------------------------------------------------------------


def merge_rows(
    existing: dict[str, Any],
    incoming: dict[str, Any],
    logger: Any,
) -> dict[str, Any]:
    """
    Merge two employee row dicts representing the same person.

    Strategy: WPS data takes precedence for identity fields (EID, passport,
    visa, bank_account, address, DOB). Cash data contributes farm_id and
    department if WPS lacks them.

    Args:
        existing: already-seen row dict
        incoming: new row dict for same person
        logger: logger instance

    Returns:
        merged row dict
    """
    merged = dict(existing)
    # WPS rows have richer identity data — prefer WPS source for identity fields
    prefer_incoming = incoming.get("payment_type") == "wps" and existing.get("payment_type") == "cash"

    for field in ("eid", "passport", "visa_no", "bank_account", "address", "dob", "nationality"):
        if not merged.get(field) and incoming.get(field):
            merged[field] = incoming[field]
        elif prefer_incoming and incoming.get(field):
            merged[field] = incoming[field]

    # Enrich from cash/Silal sheet when available
    if not merged.get("farm_id") and incoming.get("farm_id"):
        merged["farm_id"] = incoming["farm_id"]
    if not merged.get("designation") and incoming.get("designation"):
        merged["designation"] = incoming["designation"]
    if not merged.get("join_date") and incoming.get("join_date"):
        merged["join_date"] = incoming["join_date"]
    if not merged.get("department") and incoming.get("department"):
        merged["department"] = incoming["department"]

    # Keep highest salary observed (WPS data is authoritative for WPS payroll)
    if incoming.get("payment_type") == "wps":
        merged["salary"] = incoming["salary"]
        merged["net_salary"] = incoming["net_salary"]
        merged["payment_type"] = "wps"
        merged["payment_company"] = incoming["payment_company"]

    # Track merge in metadata overlay
    if "metadata_overlay" not in merged:
        merged["metadata_overlay"] = []
    merged["metadata_overlay"].append(
        {
            "source_sheet": incoming.get("source_sheet"),
            "salary": incoming.get("salary"),
            "payment_type": incoming.get("payment_type"),
        }
    )
    logger.debug(f"  Merged duplicate: {existing['full_name']} (sheets: {existing['source_sheet']} + {incoming['source_sheet']})")
    return merged


def deduplicate_rows(
    all_rows: list[dict[str, Any]],
    logger: Any,
) -> tuple[list[dict[str, Any]], int]:
    """
    Deduplicate employee rows.

    Priority order: EID (strongest) → normalized full name.

    Args:
        all_rows: combined list of rows from all sources
        logger: logger instance

    Returns:
        (deduplicated_list, merge_count)
    """
    seen_eid: dict[str, dict[str, Any]] = {}
    seen_name: dict[str, dict[str, Any]] = {}
    merge_count = 0

    for row in all_rows:
        eid = row.get("eid")
        name_key = normalize_name(row["full_name"])

        # Check by EID first
        if eid and eid in seen_eid:
            seen_eid[eid] = merge_rows(seen_eid[eid], row, logger)
            merge_count += 1
            continue

        # Check by normalized name
        if name_key in seen_name:
            merged = merge_rows(seen_name[name_key], row, logger)
            seen_name[name_key] = merged
            if eid:
                seen_eid[eid] = merged
            merge_count += 1
            continue

        # New employee
        seen_name[name_key] = row
        if eid:
            seen_eid[eid] = row

    return list(seen_name.values()), merge_count


# ---------------------------------------------------------------------------
# Pydantic validation helper
# ---------------------------------------------------------------------------


def validate_employee_doc(doc: dict[str, Any]) -> list[str]:
    """
    Validate an employee document against the Employee Pydantic model.

    Args:
        doc: employee document dict

    Returns:
        list of validation error strings (empty = valid)
    """
    try:
        # Import inline to avoid polluting module namespace
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent / "src"))
        from modules.hr.models.employee import Employee

        # Convert date objects for Pydantic
        test_doc = dict(doc)
        if isinstance(test_doc.get("hireDate"), date):
            pass  # Pydantic handles date objects natively

        # Remove non-model fields before validation
        model_fields = {
            "employeeId", "employeeCode", "firstName", "lastName",
            "arabicFirstName", "arabicMiddleName", "arabicLastName",
            "email", "phone", "department", "position", "hireDate",
            "status", "gender", "nationality", "maritalStatus",
            "emiratesId", "visaIssuancePlace", "emergencyContact",
            "divisionId", "organizationId", "createdBy", "createdAt", "updatedAt",
        }
        filtered = {k: v for k, v in test_doc.items() if k in model_fields}
        Employee(**filtered)
        return []
    except Exception as exc:
        return [str(exc)]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def run(dry_run: bool, reset: bool) -> None:
    """
    Main entry point for stage 8.

    Args:
        dry_run: if True, read and compute but do not write
        reset: if True, delete all migration-tagged docs before importing
    """
    logger = make_logger(STAGE)
    db = get_db()

    if reset:
        logger.info("[RESET] Deleting migration-tagged employees, contracts, payroll data...")
        reset_migration_data(db, RESET_COLLECTIONS, logger)

    # Ensure indexes
    if not dry_run:
        db.employees.create_index("metadata.legacyRef", unique=False, background=True)
        db.employee_contracts.create_index("metadata.legacyRef", unique=False, background=True)
        db.payroll_runs.create_index("metadata.legacyRef", unique=True, background=True)
        db.payroll_entries.create_index("metadata.legacyRef", unique=False, background=True)

    # -----------------------------------------------------------------------
    # Load all source data
    # -----------------------------------------------------------------------
    logger.info("Loading source Excel files...")

    wps1_rows = load_wps_sheet(WPS1_FILE, "PAYROLL Mar -2 2026 / Countrywise")
    logger.info(f"  WPS1 (141 expected): {len(wps1_rows)} rows loaded")

    wps2_rows = load_wps_sheet(WPS2_FILE, "PAYROLL Mar 2026 / Countrywise")
    logger.info(f"  WPS2 (1 expected): {len(wps2_rows)} rows loaded")

    cash_al_ain = load_cash_sheet_al_ain()
    logger.info(f"  Cash Al Ain: {len(cash_al_ain)} rows")

    cash_liwa = load_cash_sheet_liwa()
    logger.info(f"  Cash Liwa: {len(cash_liwa)} rows")

    cash_khaza = load_cash_sheet_khaza()
    logger.info(f"  Cash Khaza (Al Khazana): {len(cash_khaza)} rows")

    cash_al_wagen = load_cash_sheet_al_wagen()
    logger.info(f"  Cash Al Wagen: {len(cash_al_wagen)} rows")

    cash_silal = load_cash_sheet_silal_upgrade()
    logger.info(f"  Cash Silal Upgrade: {len(cash_silal)} rows")

    all_rows = wps1_rows + wps2_rows + cash_al_ain + cash_liwa + cash_khaza + cash_al_wagen + cash_silal
    logger.info(f"Total rows before dedup: {len(all_rows)}")

    # Source breakdown for summary
    source_counts = {
        "WPS1": len(wps1_rows),
        "WPS2": len(wps2_rows),
        "Cash-AlAin": len(cash_al_ain),
        "Cash-Liwa": len(cash_liwa),
        "Cash-Khaza": len(cash_khaza),
        "Cash-AlWagen": len(cash_al_wagen),
        "Cash-SilalUpgrade": len(cash_silal),
    }

    # -----------------------------------------------------------------------
    # Deduplicate
    # -----------------------------------------------------------------------
    logger.info("Deduplicating employee rows...")
    unique_rows, merge_count = deduplicate_rows(all_rows, logger)
    logger.info(f"  Unique employees after dedup: {len(unique_rows)} ({merge_count} merges performed)")

    # -----------------------------------------------------------------------
    # Payroll run ID (deterministic — stable across re-runs)
    # -----------------------------------------------------------------------
    run_id = deterministic_uuid("payroll_run", "march_2026")

    # -----------------------------------------------------------------------
    # Build and upsert documents
    # -----------------------------------------------------------------------
    emp_inserted = emp_updated = 0
    contract_inserted = contract_updated = 0
    entry_inserted = entry_updated = 0
    emp_errors: list[str] = []
    validation_errors: list[str] = []

    # Aggregation tracking for payroll run summary
    total_payroll = 0.0
    by_farm: dict[str, dict] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    by_company: dict[str, dict] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    by_payment: dict[str, dict] = defaultdict(lambda: {"count": 0, "amount": 0.0})

    # Farm name lookup (reverse: id → name) for denormalization
    farm_id_to_name: dict[str, str] = {v: k for k, v in FARM_NAME_MAP.items()}
    # Clean up duplicates in reverse map (take title-case key)
    farm_id_to_name_clean: dict[str, str] = {}
    for farm_id_val, farm_name_val in farm_id_to_name.items():
        if farm_id_val not in farm_id_to_name_clean:
            farm_id_to_name_clean[farm_id_val] = farm_name_val.title()

    logger.info(f"Processing {len(unique_rows)} unique employee records...")

    # Allocate unique employeeCode (E-NNNN) — find current max to avoid clashes
    existing_max = 0
    try:
        for doc in db.employees.find({"employeeCode": {"$regex": "^E-"}}, {"employeeCode": 1}):
            code = doc.get("employeeCode") or ""
            if code.startswith("E-"):
                try:
                    n = int(code[2:])
                    if n > existing_max: existing_max = n
                except ValueError: pass
    except Exception: pass
    emp_code_counter = existing_max

    for row in unique_rows:
        full_name = row["full_name"]
        eid = row.get("eid")
        legacy_ref = legacy_ref_for_employee(full_name, eid)

        # Build employee doc
        emp_doc = build_employee_doc(
            full_name=full_name,
            eid=eid,
            passport=row.get("passport"),
            visa_no=row.get("visa_no"),
            bank_account=row.get("bank_account"),
            dob=row.get("dob"),
            nationality=row.get("nationality"),
            address=row.get("address"),
            join_date=row.get("join_date"),
            designation=row.get("designation"),
            department=row.get("department"),
            farm_id=row.get("farm_id"),
            payment_type=row["payment_type"],
            payment_company=row.get("payment_company"),
            status_raw=row.get("status_raw"),
            source_sheet=row["source_sheet"],
        )

        # Assign unique employeeCode (E-NNNN) — reuse if legacyRef already has one
        existing = db.employees.find_one(
            {"metadata.legacyRef": emp_doc.get("metadata", {}).get("legacyRef")},
            {"employeeCode": 1}
        )
        if existing and existing.get("employeeCode"):
            emp_doc["employeeCode"] = existing["employeeCode"]
        else:
            emp_code_counter += 1
            emp_doc["employeeCode"] = f"E-{emp_code_counter:04d}"

        # Pydantic validation (proactive — catches schema drift before writes)
        v_errors = validate_employee_doc(emp_doc)
        if v_errors:
            msg = f"Validation error [{full_name}]: {v_errors[0]}"
            logger.warning(f"  WARN: {msg}")
            validation_errors.append(msg)
            # Reason: Log but continue — migration should not fail on validation
            # warnings that are non-critical (extra fields, etc.)

        emp_id = emp_doc["employeeId"]

        try:
            ins, upd = upsert_by_legacy_ref(
                db.employees, emp_doc, legacy_ref, dry_run=dry_run, logger=logger
            )
            if ins:
                emp_inserted += 1
            elif upd:
                emp_updated += 1
        except Exception as exc:
            msg = f"Employee upsert failed [{full_name}]: {exc}"
            logger.error(msg)
            emp_errors.append(msg)
            continue

        # Build contract doc
        contract_doc = build_contract_doc(
            employee_id=emp_id,
            monthly_salary=row["salary"],
            payment_type=row["payment_type"],
            payment_company=row.get("payment_company"),
            farm_id=row.get("farm_id"),
            legacy_ref=legacy_ref,
        )

        try:
            ins, upd = upsert_by_legacy_ref(
                db.employee_contracts,
                contract_doc,
                contract_doc["metadata"]["legacyRef"],
                dry_run=dry_run,
                logger=logger,
            )
            if ins:
                contract_inserted += 1
            elif upd:
                contract_updated += 1
        except Exception as exc:
            logger.error(f"Contract upsert failed [{full_name}]: {exc}")

        # Accumulate payroll run aggregates
        salary = row["salary"]
        net = row.get("net_salary", salary)
        farm_id_agg = row.get("farm_id") or "unknown"
        company_agg = row.get("payment_company") or "cash"
        pt = row["payment_type"]

        total_payroll += net
        by_farm[farm_id_agg]["count"] += 1
        by_farm[farm_id_agg]["amount"] = round(by_farm[farm_id_agg]["amount"] + net, 2)
        by_company[company_agg]["count"] += 1
        by_company[company_agg]["amount"] = round(by_company[company_agg]["amount"] + net, 2)
        by_payment[pt]["count"] += 1
        by_payment[pt]["amount"] = round(by_payment[pt]["amount"] + net, 2)

        # Build payroll entry doc
        farm_name_display = farm_id_to_name_clean.get(farm_id_agg or "", farm_id_agg or "N/A")
        entry_doc = build_payroll_entry_doc(
            run_id=run_id,
            employee_id=emp_id,
            employee_name=full_name,
            gross_salary=salary,
            deductions=row.get("deductions", 0.0),
            net_salary=net,
            farm_id=row.get("farm_id"),
            farm_name=farm_name_display if row.get("farm_id") else None,
            payment_type=pt,
            payment_company=row.get("payment_company"),
            status="paid" if net > 0 else "visa_hold",
            source_sheet=row["source_sheet"],
        )

        try:
            ins, upd = upsert_by_legacy_ref(
                db.payroll_entries,
                entry_doc,
                entry_doc["metadata"]["legacyRef"],
                dry_run=dry_run,
                logger=logger,
            )
            if ins:
                entry_inserted += 1
            elif upd:
                entry_updated += 1
        except Exception as exc:
            logger.error(f"Payroll entry upsert failed [{full_name}]: {exc}")

    # -----------------------------------------------------------------------
    # Payroll run document
    # -----------------------------------------------------------------------
    logger.info("Building payroll run summary doc...")
    run_doc = build_payroll_run_doc(
        run_id=run_id,
        total_amount=round(total_payroll, 2),
        total_employees=len(unique_rows),
        by_farm=dict(by_farm),
        by_company=dict(by_company),
        by_payment_type=dict(by_payment),
    )

    try:
        ins, upd = upsert_by_legacy_ref(
            db.payroll_runs,
            run_doc,
            run_doc["metadata"]["legacyRef"],
            dry_run=dry_run,
            logger=logger,
        )
        run_ins = 1 if ins else 0
        run_upd = 1 if upd else 0
    except Exception as exc:
        logger.error(f"Payroll run upsert failed: {exc}")
        run_ins = run_upd = 0

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    logger.info("")
    logger.info("=" * 60)
    logger.info("  SOURCE BREAKDOWN:")
    for src, cnt in source_counts.items():
        logger.info(f"    {src}: {cnt} rows")
    logger.info(f"  Dedup merges: {merge_count}")
    logger.info(f"  Unique employees: {len(unique_rows)}")
    logger.info("")
    logger.info("  PAYROLL SUMMARY (March 2026):")
    logger.info(f"    Total payroll AED: {round(total_payroll, 2):,.2f}")
    logger.info(f"    WPS employees: {by_payment.get('wps', {}).get('count', 0)}")
    logger.info(f"    Cash employees: {by_payment.get('cash', {}).get('count', 0)}")
    logger.info(f"    WPS total AED: {by_payment.get('wps', {}).get('amount', 0):,.2f}")
    logger.info(f"    Cash total AED: {by_payment.get('cash', {}).get('amount', 0):,.2f}")
    logger.info("")
    if validation_errors:
        logger.info(f"  Pydantic validation warnings: {len(validation_errors)}")
        for ve in validation_errors[:5]:
            logger.info(f"    - {ve}")
    logger.info("=" * 60)

    all_errors = emp_errors
    print_summary(
        stage=STAGE,
        rows_read=len(all_rows),
        rows_inserted=emp_inserted + contract_inserted + entry_inserted + run_ins,
        rows_updated=emp_updated + contract_updated + entry_updated + run_upd,
        rows_skipped=0,
        rows_errored=len(all_errors),
        error_samples=all_errors,
        logger=logger,
    )

    logger.info(f"  employees:          {emp_inserted} inserted, {emp_updated} updated")
    logger.info(f"  employee_contracts: {contract_inserted} inserted, {contract_updated} updated")
    logger.info(f"  payroll_entries:    {entry_inserted} inserted, {entry_updated} updated")
    logger.info(f"  payroll_runs:       {run_ins} inserted, {run_upd} updated")


if __name__ == "__main__":
    parser = make_arg_parser(
        "Stage 8: Import HR employees, contracts, and March 2026 payroll data"
    )
    args = parser.parse_args()
    run(dry_run=args.dry_run, reset=args.reset)
