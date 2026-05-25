"""
Tests for GET /reports/export/{statement}?format=pdf|xlsx (T-060.6).

Coverage:
  - 6 happy-path tests: 3 statements x 2 formats (PDF + Excel)
  - Negative: invalid format → 400
  - Negative: invalid statement slug → 400
  - Negative: missing period params for IS / CF → 400
  - Auth: non-finance role → 403
  - PDF magic bytes check (%PDF-)
  - Excel: openpyxl can open and title row / known label is present

Data seeding re-uses the helpers from test_balance_sheet.py and
test_cash_flow.py to keep things fast and isolated (SQLite in-memory).
weasyprint and openpyxl are invoked for real — this validates the full
rendering path in the test environment.
"""

import io
import uuid
from datetime import date, datetime
from decimal import Decimal

import pytest
from httpx import AsyncClient

from finance.models.orm.models import (
    AccountTypeEnum,
    CashFlowCategoryEnum,
    CompanyCode,
    DrawerEnum,
    FiscalPeriod,
    GLAccount,
    JEStatusEnum,
    JournalEntry,
    JournalEntryLine,
    PeriodStatusEnum,
)

from .conftest import auth_headers


# ---------------------------------------------------------------------------
# Shared seed helpers
# ---------------------------------------------------------------------------


async def _seed_minimal_org(db_session, org_id: str, company_code: str) -> dict:
    """
    Seed a company + a minimal CoA sufficient for all three statements.

    Returns dict of accountIds + period_id.

    Accounts seeded:
      ASSETS header + Cash leaf (CASH category) + AR leaf (WORKING_CAPITAL)
      LIABILITIES header + AP leaf (WORKING_CAPITAL)
      EQUITY header + Share Capital leaf (FINANCING)
      REVENUE leaf (P&L — for Net Income)
      OperatingCost leaf (P&L expense)
    """
    db_session.add(
        CompanyCode(
            companyCode=company_code,
            organizationId=org_id,
            legalName=f"Export Test {company_code}",
            fiscalYearStartMonth=1,
            fiscalYearStartDay=1,
            defaultCurrency="AED",
        )
    )
    period_id = str(uuid.uuid4())
    db_session.add(
        FiscalPeriod(
            periodId=period_id,
            companyCode=company_code,
            fiscalYear=2026,
            periodNumber=99,
            startDate=date(2025, 1, 1),
            endDate=date(2027, 12, 31),
            status=PeriodStatusEnum.OPEN,
        )
    )

    ids: dict = {}

    # Header accounts.
    for code, name, drawer, atype, key in [
        ("100000", "Assets Header", DrawerEnum.ASSETS, AccountTypeEnum.ASSET, "assets_hdr"),
        ("200000", "Liabilities Header", DrawerEnum.LIABILITIES, AccountTypeEnum.LIABILITY, "liab_hdr"),
        ("300000", "Equity Header", DrawerEnum.EQUITY, AccountTypeEnum.EQUITY, "equity_hdr"),
    ]:
        aid = str(uuid.uuid4())
        ids[key] = aid
        db_session.add(GLAccount(
            accountId=aid, organizationId=org_id, accountNumber=code,
            accountName=name, drawer=drawer, accountType=atype,
            parentAccountId=None, isHeader=True, isActive=True,
            cashFlowCategory=CashFlowCategoryEnum.NONE,
        ))

    # Leaf accounts.
    leaves = [
        ("110001", "Cash", DrawerEnum.ASSETS, AccountTypeEnum.ASSET,
         ids["assets_hdr"], CashFlowCategoryEnum.CASH, "cash"),
        ("121001", "Accounts Receivable", DrawerEnum.ASSETS, AccountTypeEnum.ASSET,
         ids["assets_hdr"], CashFlowCategoryEnum.WORKING_CAPITAL, "ar"),
        ("210001", "Accounts Payable", DrawerEnum.LIABILITIES, AccountTypeEnum.LIABILITY,
         ids["liab_hdr"], CashFlowCategoryEnum.WORKING_CAPITAL, "ap"),
        ("310001", "Share Capital", DrawerEnum.EQUITY, AccountTypeEnum.EQUITY,
         ids["equity_hdr"], CashFlowCategoryEnum.FINANCING, "sc"),
        ("410001", "Revenue", DrawerEnum.REVENUE, AccountTypeEnum.REVENUE,
         None, CashFlowCategoryEnum.NONE, "revenue"),
        ("510001", "Operating Cost", DrawerEnum.OPERATING_COST, AccountTypeEnum.EXPENSE,
         None, CashFlowCategoryEnum.NONE, "op_cost"),
    ]
    for code, name, drawer, atype, parent, cf_cat, key in leaves:
        aid = str(uuid.uuid4())
        ids[key] = aid
        db_session.add(GLAccount(
            accountId=aid, organizationId=org_id, accountNumber=code,
            accountName=name, drawer=drawer, accountType=atype,
            parentAccountId=parent, isHeader=False, isActive=True,
            cashFlowCategory=cf_cat,
        ))

    ids["period_id"] = period_id
    await db_session.commit()
    return ids


async def _post_je(
    db_session, org_id, company_code, lines, je_date, period_id,
):
    """Post a balanced JE (list of (accountId, debit, credit) tuples)."""
    je_id = str(uuid.uuid4())
    total_dr = sum(Decimal(str(l[1])) for l in lines)
    total_cr = sum(Decimal(str(l[2])) for l in lines)
    db_session.add(JournalEntry(
        jeId=je_id, organizationId=org_id, companyCode=company_code,
        jeNumber=f"JE-EXP-{uuid.uuid4().hex[:6].upper()}",
        jeDate=je_date, periodId=period_id,
        sourceEventType="test_export", sourceEventId=je_id,
        totalDebit=total_dr, totalCredit=total_cr,
        status=JEStatusEnum.POSTED,
        postedAt=datetime.utcnow(), postedBy="user-test",
    ))
    for i, (account_id, debit, credit) in enumerate(lines, start=1):
        db_session.add(JournalEntryLine(
            jeLineId=str(uuid.uuid4()), jeId=je_id, lineNumber=i,
            accountId=account_id,
            debit=Decimal(str(debit)),
            credit=Decimal(str(credit)),
        ))
    await db_session.commit()


# ---------------------------------------------------------------------------
# Fixtures: one seeded org per test module (function scope for isolation)
# ---------------------------------------------------------------------------


@pytest.fixture
async def seeded(db_session):
    """Provide a seeded org + posted JE for all export tests."""
    org_id = f"org-exp-{uuid.uuid4().hex[:8]}"
    company_code = f"EX{uuid.uuid4().hex[:6].upper()}"
    ids = await _seed_minimal_org(db_session, org_id, company_code)

    # Post a JE: DR Cash 10000, CR Share Capital 10000.
    await _post_je(
        db_session, org_id, company_code,
        lines=[
            (ids["cash"], "10000", "0"),
            (ids["sc"], "0", "10000"),
        ],
        je_date=date(2026, 1, 15),
        period_id=ids["period_id"],
    )
    # Post a revenue JE: DR Cash 2000, CR Revenue 2000.
    await _post_je(
        db_session, org_id, company_code,
        lines=[
            (ids["cash"], "2000", "0"),
            (ids["revenue"], "0", "2000"),
        ],
        je_date=date(2026, 3, 1),
        period_id=ids["period_id"],
    )

    return {
        "org_id": org_id,
        "company_code": company_code,
        "ids": ids,
    }


# ---------------------------------------------------------------------------
# Happy-path tests: Balance Sheet
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_balance_sheet_xlsx(client: AsyncClient, seeded) -> None:
    """
    GET /reports/export/balance-sheet?format=xlsx returns HTTP 200,
    correct Content-Type, attachment disposition, and a valid .xlsx file
    containing 'Balance Sheet' in the first cell.
    """
    import openpyxl

    resp = await client.get(
        "/api/v1/finance/reports/export/balance-sheet",
        params={
            "format": "xlsx",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "as_of_date": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    assert ".xlsx" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Row 3, col 1 should contain the statement title.
    assert ws.cell(row=3, column=1).value == "Balance Sheet"
    # Company code should appear in row 2.
    assert seeded["company_code"] in str(ws.cell(row=2, column=1).value)


@pytest.mark.asyncio
async def test_export_balance_sheet_pdf(client: AsyncClient, seeded) -> None:
    """
    GET /reports/export/balance-sheet?format=pdf returns HTTP 200,
    correct Content-Type, attachment disposition, and PDF magic bytes.
    """
    resp = await client.get(
        "/api/v1/finance/reports/export/balance-sheet",
        params={
            "format": "pdf",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "as_of_date": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert "attachment" in resp.headers["content-disposition"]
    assert ".pdf" in resp.headers["content-disposition"]
    # Validate PDF magic bytes.
    assert resp.content[:4] == b"%PDF", (
        f"Expected PDF magic bytes, got {resp.content[:8]!r}"
    )


# ---------------------------------------------------------------------------
# Happy-path tests: Income Statement
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_income_statement_xlsx(client: AsyncClient, seeded) -> None:
    """
    GET /reports/export/income-statement?format=xlsx returns 200
    with a valid .xlsx containing 'Income Statement' in the title cell.
    """
    import openpyxl

    resp = await client.get(
        "/api/v1/finance/reports/export/income-statement",
        params={
            "format": "xlsx",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "period_start": "2026-01-01",
            "period_end": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.cell(row=3, column=1).value == "Income Statement"


@pytest.mark.asyncio
async def test_export_income_statement_pdf(client: AsyncClient, seeded) -> None:
    """
    GET /reports/export/income-statement?format=pdf returns 200,
    application/pdf, and PDF magic bytes.
    """
    resp = await client.get(
        "/api/v1/finance/reports/export/income-statement",
        params={
            "format": "pdf",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "period_start": "2026-01-01",
            "period_end": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# Happy-path tests: Cash Flow
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_cash_flow_xlsx(client: AsyncClient, seeded) -> None:
    """
    GET /reports/export/cash-flow?format=xlsx returns 200 with a valid
    .xlsx file containing 'Cash Flow' in the sheet title cell.
    """
    import openpyxl

    resp = await client.get(
        "/api/v1/finance/reports/export/cash-flow",
        params={
            "format": "xlsx",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "period_start": "2026-01-01",
            "period_end": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Row 3 contains the statement title.
    assert "Cash Flow" in str(ws.cell(row=3, column=1).value)


@pytest.mark.asyncio
async def test_export_cash_flow_pdf(client: AsyncClient, seeded) -> None:
    """
    GET /reports/export/cash-flow?format=pdf returns 200,
    application/pdf, and PDF magic bytes.
    """
    resp = await client.get(
        "/api/v1/finance/reports/export/cash-flow",
        params={
            "format": "pdf",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "period_start": "2026-01-01",
            "period_end": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# Negative tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_invalid_format_returns_400(
    client: AsyncClient, seeded,
) -> None:
    """format=csv is not a supported format → HTTP 400."""
    resp = await client.get(
        "/api/v1/finance/reports/export/balance-sheet",
        params={
            "format": "csv",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "as_of_date": "2026-12-31",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 400, resp.text
    # Reason: error message must mention the invalid value to aid debugging.
    assert "csv" in resp.text.lower() or "format" in resp.text.lower()


@pytest.mark.asyncio
async def test_export_invalid_statement_returns_400(
    client: AsyncClient, seeded,
) -> None:
    """Unknown statement slug → HTTP 400."""
    resp = await client.get(
        "/api/v1/finance/reports/export/foo",
        params={
            "format": "pdf",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 400, resp.text
    assert "foo" in resp.text.lower() or "statement" in resp.text.lower()


@pytest.mark.asyncio
async def test_export_income_statement_missing_period_returns_400(
    client: AsyncClient, seeded,
) -> None:
    """income-statement without period_start / period_end → HTTP 400."""
    resp = await client.get(
        "/api/v1/finance/reports/export/income-statement",
        params={
            "format": "xlsx",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            # Intentionally omit period_start and period_end.
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 400, resp.text


@pytest.mark.asyncio
async def test_export_cash_flow_missing_period_returns_400(
    client: AsyncClient, seeded,
) -> None:
    """cash-flow without period_start / period_end → HTTP 400."""
    resp = await client.get(
        "/api/v1/finance/reports/export/cash-flow",
        params={
            "format": "pdf",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            # Intentionally omit period params.
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 400, resp.text


# ---------------------------------------------------------------------------
# Auth test
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_non_finance_role_forbidden(
    client: AsyncClient, seeded,
) -> None:
    """A role outside _READ_ROLES (e.g. 'user') receives HTTP 403."""
    resp = await client.get(
        "/api/v1/finance/reports/export/balance-sheet",
        params={
            "format": "pdf",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "as_of_date": "2026-12-31",
        },
        headers=auth_headers(role="user"),
    )
    assert resp.status_code == 403, resp.text


# ---------------------------------------------------------------------------
# Filename convention test
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_filename_contains_statement_and_date(
    client: AsyncClient, seeded,
) -> None:
    """
    Content-Disposition filename must follow the convention:
    {statement}_{period}_{company_code}.{ext}
    """
    resp = await client.get(
        "/api/v1/finance/reports/export/balance-sheet",
        params={
            "format": "xlsx",
            "organization_id": seeded["org_id"],
            "company_code": seeded["company_code"],
            "as_of_date": "2026-06-30",
        },
        headers=auth_headers(role="auditor"),
    )
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "balance-sheet" in disposition
    assert "2026-06-30" in disposition
    assert seeded["company_code"] in disposition
