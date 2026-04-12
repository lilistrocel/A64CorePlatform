"""
stage11_historical_2023_2024.py — Add FY2023 historical snapshot to revenue_reference.

Source: OldData/7-April-2026/Fully Reports - 2023-2024.xlsx
  Sheet "Expense": per-category × per-farm cost breakdown for FY2023
  Sheet "Revenue": per-farm / per-greenhouse Silal + outside-customer revenue

Both sheets produce SUMMARY-level docs only (no individual transactions).
These feed the P&L dashboard's multi-year comparison view.

All docs go into the existing `revenue_reference` collection with:
  - period.farmingYear = 2023
  - period.label = "FY2023"
  - source = "fully_reports_2023_2024_xlsx"

Run: python stage11_historical_2023_2024.py [--dry-run] [--reset]
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from common import (
    DIVISION_ID,
    MIGRATION_TAG,
    ORGANIZATION_ID,
    deterministic_uuid,
    get_db,
    make_arg_parser,
    make_logger,
    print_summary,
    reset_migration_data,
    upsert_by_legacy_ref,
    utcnow,
)

STAGE = "stage11_historical_2023_2024"

# Note: --reset targets only FY2023 docs, not the entire revenue_reference
# collection (which also contains FY2024 data from stage10).
RESET_COLLECTIONS: list[str] = []

ADMIN_UUID: str = "bff26b8f-5ce9-49b2-9126-86174eaea823"

DATA_DIR: Path = Path(__file__).parent.parent.parent.parent / "OldData" / "7-April-2026"
HISTORY_XLSX: Path = DATA_DIR / "Fully Reports - 2023-2024.xlsx"

FY2023_PERIOD: dict = {
    "start": "2023-08-01",
    "end": "2024-07-31",
    "label": "FY2023",
    "farmingYear": 2023,
}

SOURCE_TAG: str = "fully_reports_2023_2024_xlsx"

# Farm name → UUID (from migrated farms)
FARM_UUID_MAP: dict[str, str] = {
    "al ain farm": "042ab6a6-74c2-58ca-84ec-73dd3186b0d9",
    "al ain": "042ab6a6-74c2-58ca-84ec-73dd3186b0d9",
    "liwa farm 330a": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
    "liwa farm": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
    "liwa": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
    "al khazana farm": "2b34823c-ac2d-58c8-a57d-ed2c8fbbdf81",
    "al wagan farm": "b2c23ace-a0db-549c-9872-28f29658a2f3",
    "al wagen farm": "b2c23ace-a0db-549c-9872-28f29658a2f3",
    # GH / NH rows map to Liwa (they are within Liwa operation)
    "liwa gh & nh": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def safe_float(val: Any) -> Optional[float]:
    """Convert cell to float or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def resolve_farm(name: str) -> tuple[Optional[str], str]:
    """
    Resolve farm name to (farmId, displayName).

    Args:
        name: raw farm name

    Returns:
        (farmId UUID or None, display name)
    """
    key = name.strip().lower().rstrip(" ")
    return FARM_UUID_MAP.get(key), name.strip().title()


def build_ref_doc(
    legacy_ref: str,
    scope: str,
    farm_id: Optional[str],
    farm_name: Optional[str],
    channel: Optional[str],
    category: Optional[str],
    qty_kgs: Optional[float],
    amount_aed: float,
    avg_price: Optional[float],
    notes: Optional[str],
    source_row: Optional[int] = None,
) -> dict[str, Any]:
    """
    Build a revenue_reference doc for FY2023 historical data.

    Args:
        legacy_ref: stable upsert key
        scope: "expense_summary" | "revenue_summary" | "expense_category" | "revenue_detail"
        farm_id: farm UUID or None
        farm_name: farm display name
        channel: revenue channel or None
        category: category / line item label
        qty_kgs: quantity sold in kg (revenue rows only)
        amount_aed: monetary amount AED
        avg_price: average price AED/kg
        notes: free-text notes
        source_row: Excel row for traceability

    Returns:
        revenue_reference document dict
    """
    now = utcnow()
    ref_id = deterministic_uuid("revenue_ref", legacy_ref)
    return {
        "referenceId": ref_id,
        "source": SOURCE_TAG,
        "period": FY2023_PERIOD,
        "scope": scope,
        "farmId": farm_id,
        "farmName": farm_name,
        "channel": channel,
        "category": category,
        "qtyKgs": round(qty_kgs, 3) if qty_kgs is not None else None,
        "amountAed": round(amount_aed, 2),
        "avgPriceAedPerKg": round(avg_price, 4) if avg_price is not None else None,
        "notes": notes,
        "createdBy": ADMIN_UUID,
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": legacy_ref,
            "sourceRow": source_row,
        },
    }


# ---------------------------------------------------------------------------
# Expense sheet loader
# ---------------------------------------------------------------------------


def load_expense_docs() -> list[dict[str, Any]]:
    """
    Parse the Expense sheet and build expense_summary docs.

    The sheet has three sections:
      1. Purchases (rows 3-12): per-category, per-farm costs
      2. Pettycash (rows 15-28): operational petty expenses per farm
      3. Payroll (rows 31-34): salary per farm

    Column mapping (0-indexed):
      0: Category
      1: Al Ain Farm
      2: Liwa Farm 330A
      3: Al Khazana Farm
      4: Al Wagan Farm
      5: Liwa GH & NH
      6: All Farms (aggregate / indirect)
      7: Total Purchases / Total PC / Total Salary

    We emit one "expense_summary" doc per TOTAL column row (col 7),
    and one "expense_category" doc per individual category line.

    Returns:
        list of revenue_reference docs
    """
    wb = openpyxl.load_workbook(HISTORY_XLSX, data_only=True)
    ws = wb["Expense"]
    docs: list[dict[str, Any]] = []

    # Farm column mapping: col index → farm name
    FARM_COLS: dict[int, str] = {
        1: "Al Ain Farm",
        2: "Liwa Farm 330A",
        3: "Al Khazana Farm",
        4: "Al Wagan Farm",
        5: "Liwa GH & NH",
        6: "All Farms",
    }

    # Detect section headers to set context
    section_labels: dict[int, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        cell0 = str(row[0]).strip() if row[0] else ""
        if cell0.lower() in ("purchases", "pettycash", "pettycash ", "payroll"):
            section_labels[i] = cell0.strip().lower()

    # Read all data rows
    for row_idx, row_vals in enumerate(ws.iter_rows(values_only=True), 1):
        cat_raw = str(row_vals[0]).strip() if row_vals[0] else ""
        if not cat_raw:
            continue

        # Skip header and section-marker rows
        if cat_raw.lower() in (
            "purchases", "pettycash", "pettycash ", "payroll",
            "expenditure ledger",
        ):
            continue

        total_val = safe_float(row_vals[7]) if len(row_vals) > 7 else None

        # ---- Section subtotal rows (col 0 is None, col 7 has total)
        if row_vals[0] is None and total_val is not None:
            # This is a subtotal row — build an expense_summary doc
            # Find what section this belongs to (look backwards for section header)
            section_name = "expenses"
            for r in range(row_idx, 0, -1):
                if r in section_labels:
                    section_name = section_labels[r]
                    break
            legacy_ref = f"fy2023:expense_summary:{section_name}:{row_idx}"
            doc = build_ref_doc(
                legacy_ref=legacy_ref,
                scope="expense_summary",
                farm_id=None,
                farm_name=None,
                channel=None,
                category=f"{section_name.title()} Subtotal",
                qty_kgs=None,
                amount_aed=total_val,
                avg_price=None,
                notes=f"Row {row_idx} section subtotal",
                source_row=row_idx,
            )
            docs.append(doc)
            continue

        # ---- Expense category row (has total in col 7)
        if total_val is not None and total_val != 0:
            legacy_ref = f"fy2023:expense_cat:{cat_raw.lower().replace(' ', '_')}:{row_idx}"
            doc = build_ref_doc(
                legacy_ref=legacy_ref,
                scope="expense_category",
                farm_id=None,
                farm_name=None,
                channel=None,
                category=cat_raw,
                qty_kgs=None,
                amount_aed=total_val,
                avg_price=None,
                notes=None,
                source_row=row_idx,
            )
            docs.append(doc)

            # Also emit per-farm docs for non-zero values
            for col_idx, farm_name in FARM_COLS.items():
                farm_val = safe_float(row_vals[col_idx]) if col_idx < len(row_vals) else None
                if farm_val and farm_val != 0:
                    farm_id, farm_display = resolve_farm(farm_name)
                    farm_legacy_ref = (
                        f"fy2023:expense_farm:{cat_raw.lower().replace(' ', '_')}:"
                        f"{farm_name.lower().replace(' ', '_')}"
                    )
                    farm_doc = build_ref_doc(
                        legacy_ref=farm_legacy_ref,
                        scope="expense_category",
                        farm_id=farm_id,
                        farm_name=farm_display,
                        channel=None,
                        category=cat_raw,
                        qty_kgs=None,
                        amount_aed=farm_val,
                        avg_price=None,
                        notes=None,
                        source_row=row_idx,
                    )
                    docs.append(farm_doc)

    return docs


# ---------------------------------------------------------------------------
# Revenue sheet loader
# ---------------------------------------------------------------------------


def load_revenue_docs() -> list[dict[str, Any]]:
    """
    Parse the Revenue sheet and build revenue_summary docs.

    Sheet has two sections:
      1. Silal Sales (rows 3-15): per-farm/greenhouse Silal channel
      2. Outside Customer Sales (rows 17+): per-customer outside sales

    Column mapping:
      0: Inventory Sources / Customer name
      1: Supplier Code
      2: Sold Qty - Kgs
      3: Sales Revenue - AED
      4: Avg Price

    Returns:
        list of revenue_reference docs
    """
    wb = openpyxl.load_workbook(HISTORY_XLSX, data_only=True)
    ws = wb["Revenue"]
    docs: list[dict[str, Any]] = []

    current_channel = "silal"

    for row_idx, row_vals in enumerate(ws.iter_rows(values_only=True), 1):
        source_col = str(row_vals[0]).strip() if row_vals[0] else ""

        # Section header detection
        if "outside customer" in source_col.lower():
            current_channel = "market_outside"
            continue
        if source_col.lower() in ("inventory sources", ""):
            continue
        if "silal sales" in source_col.lower():
            current_channel = "silal"
            continue

        qty = safe_float(row_vals[2]) if len(row_vals) > 2 else None
        revenue = safe_float(row_vals[3]) if len(row_vals) > 3 else None
        avg = safe_float(row_vals[4]) if len(row_vals) > 4 else None

        if revenue is None or revenue == 0:
            continue

        # Skip subtotal / total rows (no Inventory Source name, or "Total" keyword)
        if source_col.lower().startswith("total") or source_col.lower().startswith("silal sales"):
            continue
        # Skip blank / numeric-only source names (subtotal rows have None in col 0)
        if row_vals[0] is None:
            continue

        # Try to resolve as farm, otherwise treat as customer/greenhouse
        farm_id, farm_display = resolve_farm(source_col)
        # Green Nation, Al Hesrh etc. are customers, not farms
        is_customer = not farm_id and current_channel == "market_outside"

        legacy_ref = (
            f"fy2023:revenue:{current_channel}:"
            f"{source_col.lower().replace(' ', '_')[:50]}"
        )
        scope = "revenue_summary"

        doc = build_ref_doc(
            legacy_ref=legacy_ref,
            scope=scope,
            farm_id=farm_id,
            farm_name=farm_display if farm_id else None,
            channel=current_channel,
            category=source_col if is_customer else None,
            qty_kgs=qty,
            amount_aed=revenue,
            avg_price=avg,
            notes=f"Customer: {source_col}" if is_customer else None,
            source_row=row_idx,
        )
        docs.append(doc)

    return docs


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def run(dry_run: bool, reset: bool) -> None:
    """
    Main entry point for stage 11.

    Args:
        dry_run: if True, read and compute but do not write
        reset: if True, delete only FY2023 docs from revenue_reference
    """
    logger = make_logger(STAGE)
    db = get_db()

    if reset:
        # Reason: targeted delete — only FY2023 stage11 docs
        result = db.revenue_reference.delete_many(
            {
                "metadata.migratedFrom": MIGRATION_TAG,
                "source": SOURCE_TAG,
            }
        )
        logger.info(f"[RESET] Deleted {result.deleted_count} FY2023 revenue_reference docs")

    if not dry_run:
        db.revenue_reference.create_index("metadata.legacyRef", unique=False, background=True)

    # -----------------------------------------------------------------------
    # Load docs
    # -----------------------------------------------------------------------
    logger.info("Loading FY2023 expense data from Expense sheet...")
    expense_docs = load_expense_docs()
    logger.info(f"  Expense docs: {len(expense_docs)}")

    logger.info("Loading FY2023 revenue data from Revenue sheet...")
    revenue_docs = load_revenue_docs()
    logger.info(f"  Revenue docs: {len(revenue_docs)}")

    all_docs = expense_docs + revenue_docs
    logger.info(f"Total FY2023 docs to upsert: {len(all_docs)}")

    if dry_run:
        logger.info("[DRY-RUN] Sample docs:")
        for doc in all_docs[:10]:
            logger.info(
                f"  scope={doc['scope']} farm={doc.get('farmName')} "
                f"channel={doc.get('channel')} cat={doc.get('category')} "
                f"AED={doc['amountAed']:,.2f}"
            )

    # -----------------------------------------------------------------------
    # Upsert
    # -----------------------------------------------------------------------
    inserted = updated = 0
    error_samples: list[str] = []

    for doc in all_docs:
        legacy_ref = doc["metadata"]["legacyRef"]
        try:
            ins, upd = upsert_by_legacy_ref(
                db.revenue_reference, doc, legacy_ref, dry_run=dry_run, logger=logger
            )
            if ins:
                inserted += 1
            elif upd:
                updated += 1
        except Exception as exc:
            msg = f"Upsert failed [{legacy_ref}]: {exc}"
            logger.error(msg)
            error_samples.append(msg)

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    total_expense_aed = sum(
        d["amountAed"]
        for d in expense_docs
        if d["scope"] == "expense_summary" and "subtotal" in (d.get("category") or "").lower()
    )
    total_revenue_aed = sum(
        d["amountAed"]
        for d in revenue_docs
    )
    logger.info("")
    logger.info("=" * 60)
    logger.info("  HISTORICAL FY2023 SUMMARY:")
    logger.info(f"    Expense docs: {len(expense_docs)}")
    logger.info(f"    Revenue docs: {len(revenue_docs)}")
    logger.info(f"    Total revenue AED: {total_revenue_aed:,.2f}")
    logger.info("=" * 60)

    print_summary(
        stage=STAGE,
        rows_read=len(all_docs),
        rows_inserted=inserted,
        rows_updated=updated,
        rows_skipped=0,
        rows_errored=len(error_samples),
        error_samples=error_samples,
        logger=logger,
    )


if __name__ == "__main__":
    parser = make_arg_parser(
        "Stage 11: Add FY2023 historical expense/revenue snapshot to revenue_reference"
    )
    args = parser.parse_args()
    run(dry_run=args.dry_run, reset=args.reset)
