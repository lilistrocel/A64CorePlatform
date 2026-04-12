"""
stage6_purchase_register.py — Parse Purchase Register Excel into purchase_register collection.

Source: OldData/7-April-2026/Purchase Register 2025-2026.xlsx
Sheets: "AGRI NOVA INVESTMENT - L.L.C" and "AL SAMARAT TRADING L.L.C"
        (treated as ONE pooled buyer entity per user decision)

Tally hierarchical format state machine:
  Voucher header row: col[0]=date, col[1]=supplier, col[4]=vch_type, col[5]=vch_no, col[7]=credit
  "Purchases" / "Purchase - Raw Materials" sub-row: col[1] starts with "Purchase", col[6]=subtotal
  Item row: col[0]=None, col[1]=item_name, col[2]=qty, col[3]=rate, col[4]=amount
  "VAT5%" row: col[1]=="VAT5%" or "VAT 5%", col[6]=vat_amount

New collection: purchase_register

Document shape:
  { voucherId, date, supplier, vchType, vchNo, totalAmount, vatAmount,
    items: [{name, qty, rate, amount, mappedCropName}],
    buyerEntity: "agri_nova"|"al_samarat",
    source: "purchase_register_2025_2026",
    metadata: { migratedFrom, legacyRef, sheetName } }

Crop mapping (best-effort, no failure):
  - Item name contains "Seeds" → try to extract crop name after "Seeds #" or "Seeds -"
  - Item name contains a known crop name substring → map it
  - Unmatched → mappedCropName=None, no failure

Run:  python stage6_purchase_register.py [--dry-run] [--reset]
"""

from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

sys.path.insert(0, str(Path(__file__).parent))

from common import (
    DIVISION_ID,
    MIGRATION_TAG,
    ORGANIZATION_ID,
    PURCHASE_EXCEL,
    deterministic_uuid,
    get_db,
    load_crop_map,
    make_arg_parser,
    make_logger,
    print_summary,
    reset_migration_data,
    utcnow,
)

STAGE = "stage6_purchase_register"
RESET_COLLECTIONS = ["purchase_register"]

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SHEET_TO_ENTITY = {
    "AGRI NOVA INVESTMENT - L.L.C": "agri_nova",
    "AL SAMARAT TRADING L.L.C": "al_samarat",
}


# ---------------------------------------------------------------------------
# Tally sheet parser
# ---------------------------------------------------------------------------

class VoucherParseError(Exception):
    """Raised when a voucher row cannot be parsed."""


def _is_date(val: Any) -> bool:
    """Check if a value is a datetime (from openpyxl)."""
    return isinstance(val, datetime)


def _safe_float(val: Any) -> Optional[float]:
    """Safely convert a cell value to float."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _is_vat_row(row: tuple) -> bool:
    """Detect a VAT row: col[1] starts with 'VAT'."""
    return (
        row[1] is not None
        and str(row[1]).strip().upper().startswith("VAT")
    )


def _is_purchase_subrow(row: tuple) -> bool:
    """Detect a 'Purchases' / 'Purchase - Raw Materials' subtotal row."""
    if row[1] is None:
        return False
    s = str(row[1]).strip().lower()
    return s.startswith("purchase")


def _is_item_row(row: tuple) -> bool:
    """Detect an item line: col[0] is None, col[1] is string, col[4] looks like an amount."""
    return (
        row[0] is None
        and row[1] is not None
        and str(row[1]).strip() != ""
        and not _is_vat_row(row)
        and not _is_purchase_subrow(row)
        and _safe_float(row[4]) is not None
    )


def _is_voucher_header(row: tuple) -> bool:
    """Detect a voucher header row: col[0] is a datetime."""
    return _is_date(row[0]) and row[1] is not None and str(row[1]).strip() != ""


def _extract_unit_from_format(fmt: Optional[str]) -> Optional[str]:
    """
    Extract unit string from an Excel number_format like '""0" BAG"' or '""0.00"/Kg"'.
    Returns unit name ('BAG', 'Kg', 'Ltr', etc.) or None.
    """
    if not fmt or fmt in ("General", "@"):
        return None
    import re
    cleaned = re.sub(r'["0.,#/]+', " ", fmt).strip()
    return cleaned if cleaned else None


def parse_sheet(ws) -> list[dict]:
    """
    Parse a single Tally-format worksheet into voucher dicts.

    State machine:
      IDLE → (voucher header) → IN_VOUCHER → (item lines) → (VAT row) → IDLE

    Args:
        ws: openpyxl worksheet (read_only)

    Returns:
        list of voucher dicts with items list
    """
    vouchers = []
    current_voucher: Optional[dict] = None
    # Data starts at row 6 in AGRI NOVA, row 7 in AL SAMARAT (slight variation)
    # We handle this by scanning all rows and using the state machine

    for row_cells in ws.iter_rows(min_row=4, max_col=8):
        # Pad to 8 cells
        cells = list(row_cells) + [None] * max(0, 8 - len(row_cells))
        row = tuple(c.value if c is not None else None for c in cells[:8])
        qty_cell = cells[2]
        rate_cell = cells[3]

        if _is_voucher_header(row):
            # Save previous voucher if any
            if current_voucher:
                vouchers.append(current_voucher)
            # Start new voucher
            current_voucher = {
                "date": row[0],
                "supplier": str(row[1]).strip(),
                "vchType": str(row[4]).strip() if row[4] else None,
                "vchNo": str(row[5]).strip() if row[5] else None,
                "totalAmount": _safe_float(row[7]),
                "vatAmount": None,
                "purchasesSubtotal": None,
                "items": [],
            }

        elif current_voucher is not None:
            if _is_purchase_subrow(row):
                current_voucher["purchasesSubtotal"] = _safe_float(row[6])

            elif _is_vat_row(row):
                vat = _safe_float(row[6])
                if current_voucher["vatAmount"] is None:
                    current_voucher["vatAmount"] = vat
                else:
                    # Some vouchers have multiple VAT rows — sum them
                    current_voucher["vatAmount"] = (current_voucher["vatAmount"] or 0) + (vat or 0)

            elif _is_item_row(row):
                item_name = str(row[1]).strip()
                qty = _safe_float(row[2])
                rate = _safe_float(row[3])
                amount = _safe_float(row[4])
                # Extract unit from cell number format
                unit = None
                if qty_cell is not None:
                    unit = _extract_unit_from_format(getattr(qty_cell, "number_format", None))
                if not unit and rate_cell is not None:
                    unit = _extract_unit_from_format(getattr(rate_cell, "number_format", None))
                current_voucher["items"].append({
                    "name": item_name,
                    "qty": qty,
                    "unit": unit,
                    "rate": rate,
                    "amount": amount,
                })

    # Don't forget the last voucher
    if current_voucher:
        vouchers.append(current_voucher)

    return vouchers


# ---------------------------------------------------------------------------
# Crop mapping for purchase items
# ---------------------------------------------------------------------------


def map_item_to_crop(item_name: str, crop_names_lower: set[str]) -> Optional[str]:
    """
    Attempt to map a purchase item name to a known crop name.

    Strategies:
    1. "Seeds # CropName" or "Seeds - CropName" → extract after delimiter
    2. Any known crop name appears as substring in item_name

    Args:
        item_name: raw item name from purchase register
        crop_names_lower: set of lowercased known crop names

    Returns:
        matched crop name (original case from item after delimiter) or None
    """
    lower = item_name.strip().lower()

    # Strategy 1: seeds prefix
    seeds_match = re.match(r"seeds?\s*[#\-:]\s*(.+)", lower)
    if seeds_match:
        candidate = seeds_match.group(1).strip()
        if candidate in crop_names_lower:
            return candidate.title()
        # Partial match
        for crop in crop_names_lower:
            if crop in candidate or candidate in crop:
                return crop.title()

    # Strategy 2: any crop name is substring
    for crop in crop_names_lower:
        if len(crop) > 4 and crop in lower:
            return crop.title()

    return None


# ---------------------------------------------------------------------------
# Document builder
# ---------------------------------------------------------------------------

def build_voucher_doc(
    voucher: dict,
    buyer_entity: str,
    sheet_name: str,
    crop_names_lower: set[str],
) -> dict:
    """
    Build a purchase_register document from a parsed voucher.

    Args:
        voucher: from parse_sheet()
        buyer_entity: "agri_nova" or "al_samarat"
        sheet_name: original sheet name
        crop_names_lower: set of lowercased crop names for item mapping

    Returns:
        purchase_register document dict
    """
    now = utcnow()

    # Deterministic key: buyer + vchNo + date
    date_str = voucher["date"].strftime("%Y-%m-%d") if isinstance(voucher["date"], datetime) else str(voucher["date"])
    legacy_key = f"{buyer_entity}|{voucher.get('vchNo','?')}|{date_str}|{voucher.get('supplier','?')}"
    voucher_uuid = deterministic_uuid("purchase_voucher", legacy_key)

    # Enrich items with crop mapping
    enriched_items = []
    for item in voucher.get("items", []):
        mapped_crop = map_item_to_crop(item["name"], crop_names_lower)
        enriched_items.append({
            "name": item["name"],
            "qty": item["qty"],
            "unit": item.get("unit"),
            "rate": item["rate"],
            "amount": item["amount"],
            "mappedCropName": mapped_crop,
        })

    return {
        "voucherId": voucher_uuid,
        "date": voucher["date"],
        "supplier": voucher.get("supplier"),
        "vchType": voucher.get("vchType"),
        "vchNo": voucher.get("vchNo"),
        "totalAmount": voucher.get("totalAmount"),
        "purchasesSubtotal": voucher.get("purchasesSubtotal"),
        "vatAmount": voucher.get("vatAmount"),
        "items": enriched_items,
        "buyerEntity": buyer_entity,
        "source": "purchase_register_2025_2026",
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": voucher["date"] if isinstance(voucher["date"], datetime) else now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": voucher_uuid,
            "sheetName": sheet_name,
        },
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def run(dry_run: bool, reset: bool) -> None:
    """
    Main entry point for stage 6.

    Args:
        dry_run: if True, read and compute but do not write
        reset: if True, delete all migration-tagged purchase_register docs
    """
    logger = make_logger(STAGE)
    db = get_db()

    if reset:
        logger.info("[RESET] Deleting migration-tagged purchase_register docs...")
        reset_migration_data(db, RESET_COLLECTIONS, logger)

    # Ensure index
    if not dry_run:
        db.purchase_register.create_index("metadata.legacyRef", background=True)
        db.purchase_register.create_index("date", background=True)

    # Load crop map for item matching
    logger.info("Loading crop map...")
    crop_map = load_crop_map(db)
    crop_names_lower = set(crop_map.keys())
    logger.info(f"Crop map: {len(crop_names_lower)} crops")

    logger.info(f"Opening {PURCHASE_EXCEL}...")
    # Load with read_only=False so cell.number_format is accessible for unit extraction
    wb = openpyxl.load_workbook(PURCHASE_EXCEL, data_only=True)

    inserted = updated = skipped = 0
    error_samples: list[str] = []
    total_vouchers = 0
    total_items = 0

    for sheet_name in wb.sheetnames:
        entity = SHEET_TO_ENTITY.get(sheet_name)
        if not entity:
            logger.warning(f"Skipping unknown sheet: {sheet_name}")
            continue

        logger.info(f"Parsing sheet: {sheet_name} → entity={entity}")
        ws = wb[sheet_name]
        vouchers = parse_sheet(ws)
        logger.info(f"  Parsed {len(vouchers)} vouchers from {sheet_name}")

        for voucher in vouchers:
            doc = build_voucher_doc(voucher, entity, sheet_name, crop_names_lower)
            legacy_ref = doc["metadata"]["legacyRef"]
            total_vouchers += 1
            total_items += len(doc["items"])

            try:
                if not dry_run:
                    result = db.purchase_register.replace_one(
                        {"metadata.legacyRef": legacy_ref},
                        doc,
                        upsert=True,
                    )
                    if result.upserted_id:
                        inserted += 1
                    elif result.modified_count > 0:
                        updated += 1
                else:
                    logger.debug(
                        f"  [DRY-RUN] Voucher {voucher.get('vchNo')} "
                        f"supplier={voucher.get('supplier','?')[:30]} "
                        f"items={len(voucher['items'])} "
                        f"total={voucher.get('totalAmount')}"
                    )
                    inserted += 1  # Count as "would insert" for dry-run reporting
            except Exception as exc:
                msg = f"Voucher upsert failed key={legacy_ref}: {exc}"
                logger.error(msg)
                error_samples.append(msg)

    wb.close()

    logger.info(f"Total vouchers processed: {total_vouchers}")
    logger.info(f"Total line items: {total_items}")

    # Report crop-mapped items
    mapped_items_count = 0
    if not dry_run:
        for doc in db.purchase_register.find(
            {"metadata.migratedFrom": MIGRATION_TAG},
            {"items": 1, "_id": 0},
        ):
            mapped_items_count += sum(
                1 for item in doc.get("items", []) if item.get("mappedCropName")
            )
    logger.info(f"Purchase items mapped to crop: {mapped_items_count}")

    print_summary(
        stage=STAGE,
        rows_read=total_vouchers,
        rows_inserted=inserted,
        rows_updated=updated,
        rows_skipped=skipped,
        rows_errored=len(error_samples),
        error_samples=error_samples,
        logger=logger,
    )

    if error_samples:
        logger.warning(f"Stage completed with {len(error_samples)} error(s). Review log.")
        sys.exit(1)


if __name__ == "__main__":
    parser = make_arg_parser(
        "Stage 6: Parse Purchase Register Excel → purchase_register collection"
    )
    args = parser.parse_args()
    run(dry_run=args.dry_run, reset=args.reset)
