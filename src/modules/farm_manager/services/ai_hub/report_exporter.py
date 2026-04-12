"""
AI Hub - Report Exporter

Converts AI-generated markdown report text to downloadable PDF or Excel bytes.
Both methods are pure synchronous utilities; they do not touch the database or
any external service, so they are safe to call directly from async endpoints
(output sizes are small enough that blocking is not a concern).

Supported formats:
  - PDF  via reportlab (A4, styled headings, tables, bullet points)
  - XLSX via openpyxl  (single sheet, styled headings, tables)
"""

import io
import re
from datetime import datetime, timezone
from typing import List


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _strip_bold_italic(text: str) -> str:
    """Remove markdown bold/italic markers, returning plain text."""
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    text = re.sub(r"\*([^*]+)\*", r"\1", text)
    return text


def _to_reportlab_inline(text: str) -> str:
    """Convert markdown bold/italic to ReportLab XML tags."""
    text = re.sub(r"\*\*([^*]+)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"\*([^*]+)\*", r"<i>\1</i>", text)
    return text


def _parse_markdown_table(lines: List[str], start: int) -> tuple[List[List[str]], int]:
    """
    Parse a markdown pipe table beginning at `start`.

    Args:
        lines: All lines of the markdown document.
        start: Index of the header row (the line containing '|').

    Returns:
        A tuple of (table_rows, next_line_index) where table_rows contains
        only non-separator rows with bold markers stripped, and next_line_index
        is the index of the first line after the table.
    """
    table_rows: List[List[str]] = []
    i = start
    while i < len(lines) and "|" in lines[i]:
        cells = [c.strip() for c in lines[i].split("|") if c.strip()]
        # Reason: Skip separator rows (e.g. |---|---|)
        if not all(set(c.strip()) <= set("-: ") for c in cells):
            cells = [_strip_bold_italic(c) for c in cells]
            table_rows.append(cells)
        i += 1
    return table_rows, i


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class ReportExporter:
    """Converts markdown report text to downloadable PDF or Excel format."""

    @staticmethod
    def markdown_to_pdf(markdown_text: str, title: str = "AI Hub Report") -> bytes:
        """
        Convert a markdown report to PDF bytes (A4 page size).

        Args:
            markdown_text: The full markdown content produced by the AI.
            title:         Document title shown at the top and in metadata.

        Returns:
            Raw PDF bytes suitable for a FileResponse or Response body.

        Raises:
            ImportError: If reportlab is not installed.
            Exception:   Propagated from reportlab if document build fails.
        """
        # Reason: Deferred import — reportlab is optional at module load time
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=20 * mm,
            rightMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
            title=title,
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontSize=20,
            textColor=HexColor("#1976d2"),
            spaceAfter=12,
        )
        date_style = ParagraphStyle(
            "ReportDate",
            parent=styles["Normal"],
            fontSize=9,
            textColor=HexColor("#999999"),
        )
        h2_style = ParagraphStyle(
            "H2",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=HexColor("#333333"),
            spaceBefore=12,
            spaceAfter=6,
        )
        h3_style = ParagraphStyle(
            "H3",
            parent=styles["Heading3"],
            fontSize=12,
            textColor=HexColor("#555555"),
            spaceBefore=8,
            spaceAfter=4,
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            spaceAfter=4,
        )
        bullet_style = ParagraphStyle(
            "Bullet",
            parent=body_style,
            leftIndent=20,
            bulletIndent=10,
        )

        elements = []

        # Header block
        elements.append(Paragraph(title, title_style))
        generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        elements.append(Paragraph(f"Generated: {generated_at}", date_style))
        elements.append(Spacer(1, 10 * mm))
        elements.append(HRFlowable(width="100%", color=HexColor("#e0e0e0")))
        elements.append(Spacer(1, 5 * mm))

        lines = markdown_text.split("\n")
        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if not line:
                elements.append(Spacer(1, 3 * mm))
                i += 1
                continue

            # Headings
            if line.startswith("### "):
                elements.append(Paragraph(line[4:], h3_style))
            elif line.startswith("## "):
                elements.append(Paragraph(line[3:], h2_style))
            elif line.startswith("# "):
                elements.append(Paragraph(line[2:], h2_style))

            # Horizontal rule
            elif line.startswith("---"):
                elements.append(Spacer(1, 3 * mm))
                elements.append(HRFlowable(width="100%", color=HexColor("#e0e0e0")))
                elements.append(Spacer(1, 3 * mm))

            # Markdown pipe table — detected by '|' on current line and '---' on next
            elif "|" in line and i + 1 < len(lines) and "---" in lines[i + 1]:
                table_rows, i = _parse_markdown_table(lines, i)
                if table_rows:
                    # Reason: Normalize column count so Table() doesn't raise
                    max_cols = max(len(r) for r in table_rows)
                    for row in table_rows:
                        while len(row) < max_cols:
                            row.append("")

                    t = Table(table_rows)
                    t.setStyle(
                        TableStyle(
                            [
                                (
                                    "BACKGROUND",
                                    (0, 0),
                                    (-1, 0),
                                    HexColor("#f5f5f5"),
                                ),
                                (
                                    "TEXTCOLOR",
                                    (0, 0),
                                    (-1, 0),
                                    HexColor("#333333"),
                                ),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, -1), 9),
                                (
                                    "GRID",
                                    (0, 0),
                                    (-1, -1),
                                    0.5,
                                    HexColor("#e0e0e0"),
                                ),
                                ("PADDING", (0, 0), (-1, -1), 6),
                                (
                                    "ROWBACKGROUNDS",
                                    (0, 1),
                                    (-1, -1),
                                    [HexColor("#ffffff"), HexColor("#fafafa")],
                                ),
                            ]
                        )
                    )
                    elements.append(Spacer(1, 3 * mm))
                    elements.append(t)
                    elements.append(Spacer(1, 3 * mm))
                continue  # `i` already advanced by _parse_markdown_table

            # Unordered bullet
            elif line.startswith("- ") or line.startswith("* "):
                text = _to_reportlab_inline(line[2:])
                elements.append(Paragraph(f"• {text}", bullet_style))

            # Ordered list
            elif re.match(r"^\d+\.\s", line):
                text = _to_reportlab_inline(re.sub(r"^\d+\.\s", "", line))
                elements.append(Paragraph(text, bullet_style))

            # Regular paragraph
            else:
                text = _to_reportlab_inline(line)
                elements.append(Paragraph(text, body_style))

            i += 1

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def markdown_to_excel(markdown_text: str, title: str = "AI Hub Report") -> bytes:
        """
        Convert a markdown report to Excel (.xlsx) bytes.

        Headings become bold cells; pipe tables become styled worksheet tables;
        bullet/numbered list items are indented. One sheet named "Report".

        Args:
            markdown_text: The full markdown content produced by the AI.
            title:         Workbook title shown in cell A1.

        Returns:
            Raw XLSX bytes suitable for a FileResponse or Response body.

        Raises:
            ImportError: If openpyxl is not installed.
            Exception:   Propagated from openpyxl if workbook build fails.
        """
        # Reason: Deferred import — openpyxl is optional at module load time
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # ---- Style definitions -----------------------------------------------
        title_font = Font(name="Calibri", size=16, bold=True, color="1976D2")
        date_font = Font(name="Calibri", size=9, color="999999")
        heading_font = Font(name="Calibri", size=12, bold=True, color="333333")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1976D2", end_color="1976D2", fill_type="solid"
        )
        body_font = Font(name="Calibri", size=10)
        thin = Side(style="thin", color="E0E0E0")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ---- Header rows -----------------------------------------------------
        row = 1
        ws.cell(row=row, column=1, value=title).font = title_font
        row += 1
        generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        ws.cell(row=row, column=1, value=f"Generated: {generated_at}").font = date_font
        row += 2

        # ---- Body ------------------------------------------------------------
        lines = markdown_text.split("\n")
        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if not line:
                row += 1
                i += 1
                continue

            clean = _strip_bold_italic(line)

            if line.startswith("# "):
                ws.cell(row=row, column=1, value=clean[2:]).font = title_font
                row += 1

            elif line.startswith("## ") or line.startswith("### "):
                prefix = 3 if line.startswith("## ") else 4
                ws.cell(row=row, column=1, value=clean[prefix:]).font = heading_font
                row += 1

            elif line.startswith("---"):
                # Reason: Represent horizontal rules as blank rows for readability
                row += 1
                i += 1
                continue

            # Markdown pipe table
            elif "|" in line and i + 1 < len(lines) and "---" in lines[i + 1]:
                table_rows, i = _parse_markdown_table(lines, i)
                if table_rows:
                    # Header row
                    for col_idx, val in enumerate(table_rows[0], 1):
                        cell = ws.cell(row=row, column=col_idx, value=val)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border
                    row += 1

                    # Data rows
                    for data_row in table_rows[1:]:
                        for col_idx, val in enumerate(data_row, 1):
                            cell = ws.cell(row=row, column=col_idx, value=val)
                            cell.font = body_font
                            cell.border = thin_border
                        row += 1
                    row += 1  # blank row after table
                continue  # `i` already advanced

            # Unordered bullet
            elif line.startswith("- ") or line.startswith("* "):
                ws.cell(row=row, column=1, value=f"  \u2022 {clean[2:]}").font = body_font
                row += 1

            # Ordered list
            elif re.match(r"^\d+\.\s", line):
                ws.cell(
                    row=row,
                    column=1,
                    value=f"  {re.sub(r'^\\d+\\.\\s', '', clean)}",
                ).font = body_font
                row += 1

            else:
                ws.cell(row=row, column=1, value=clean).font = body_font
                row += 1

            i += 1

        # Auto-fit column widths (capped at 80 to prevent absurdly wide columns)
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 80)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
