"""
Excel Handler

openpyxl-based export and import for the Fertilizer Cost Calculator.

Export:
  export_calculation(response) → bytes (.xlsx)
  - "Calculation" sheet: per-crop blocks with ingredient rows + subtotals
  - "Warnings" sheet: if any warnings exist

Import:
  import_crops(file_bytes) → ParsedImport
  - Reads first sheet, expects "Crop Name", "Points", and optionally "Net Yield (kg)"
    header columns (case-insensitive).
  - If "Net Yield (kg)" is present and positive for a row, it takes precedence over
    "Points": dripper points are computed from the plant's yieldInfo.
  - Returns parsed items + skipped rows + warnings.

Template:
  build_import_template() → bytes (.xlsx)
  - Generates a sample file with three columns: Crop Name | Points | Net Yield (kg)
  - Includes two example rows demonstrating both input modes and one explanatory note row.
"""

import math
import re
from io import BytesIO
from datetime import date
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ...services.database import farm_db
from ...models.tools.calculator_request import (
    CalculateResponse,
    ParsedImport,
    ParsedImportItem,
    SkippedRow,
)

# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
_HEADER_BG = "1F6AA5"   # dark blue
_CROP_BG = "BDD7EE"     # light blue
_TOTAL_BG = "FFF2CC"    # light yellow
_GRAND_BG = "FCE4D6"    # light orange
_NOTE_FONT_COLOR = "888888"  # light grey

# ---------------------------------------------------------------------------
# Excel number formats — comma-separated thousands, up to 2 decimals, no
# trailing zeros. Matches the UI's `fmtUpTo2`/`fmtYield` display.
# ---------------------------------------------------------------------------
_FMT_INT = "#,##0"
_FMT_DECIMAL = "#,##0.##"

# ---------------------------------------------------------------------------
# Maximum allowed points value (mirrors CalculateItem.points upper bound)
# ---------------------------------------------------------------------------
_MAX_POINTS = 10_000_000

# ---------------------------------------------------------------------------
# Regex: matches "net yield" with optional trailing "(kg)" / "(lbs)" etc.
# Used for case-insensitive header detection.
# ---------------------------------------------------------------------------
_NET_YIELD_PATTERN = re.compile(r"^net\s+yield(\s*\(.*\))?$", re.IGNORECASE)


def build_import_template() -> bytes:
    """
    Produce a sample .xlsx file users can fill in and re-upload via /import.

    Layout (sheet "Crops"):
    - Row 1: header — "Crop Name" | "Points" | "Net Yield (kg)"
    - Row 2: example row with Points filled, Net Yield empty.
    - Row 3: example row with Net Yield filled, Points empty.
    - Row 4: italic instruction note (light grey) explaining the two modes.

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Crops"

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18

    # -- Header row --
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    for col, label in enumerate(("Crop Name", "Points", "Net Yield (kg)"), start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # -- Row 2: Points-mode example --
    ws.cell(
        row=2,
        column=1,
        value="Replace with crop name (must match Plant Library)",
    )
    ws.cell(row=2, column=2, value=100)
    # Column C intentionally left empty

    # -- Row 3: Net-Yield-mode example --
    ws.cell(row=3, column=1, value="e.g. Potato")
    # Column B intentionally left empty
    ws.cell(row=3, column=3, value=500)

    # -- Row 4: explanatory note --
    note_font = Font(italic=True, size=9, color=_NOTE_FONT_COLOR)
    note_cell = ws.cell(
        row=4,
        column=1,
        value="Fill Points OR Net Yield (kg). If both, Net Yield wins.",
    )
    note_cell.font = note_font

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_calculation(
    response: CalculateResponse,
    yield_info_by_plant: Optional[dict] = None,
) -> bytes:
    """
    Produce a .xlsx file from a CalculateResponse.

    Workbook layout:
    - Sheet "Per Crop": one block per crop with ingredient rows + subtotal,
      grand total row at the bottom plus a "TOTAL YIELD" row when yield info
      is available.
    - Sheet "Per Input": one row per chemical aggregated across all crops with
      total qty + total cost, grand total row at the bottom.
    - Sheet "Warnings": one warning per row, only if warnings exist.

    Args:
        response: CalculateResponse from the calculator engine.
        yield_info_by_plant: Optional dict mapping plantDataId (str) → yieldInfo
            dict ({yieldPerPlant, yieldUnit, seedsPerPlantingPoint,
            expectedWastePercentage}). When provided, an "Est. Yield" column
            is added to Per Crop and a TOTAL YIELD row appears alongside the
            grand total cost. When None or empty, the sheet renders without
            yield columns (back-compat).

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Per Crop"

    yield_info_by_plant = yield_info_by_plant or {}
    include_yield = bool(yield_info_by_plant)

    # --- Column widths ---
    # Layout when include_yield: Crop | Points | Cycle Days | Est. Yield | Chemical | Qty | Unit | Unit Price | Total Cost
    # Layout without:           Crop | Points | Cycle Days | Chemical | Qty | Unit | Unit Price | Total Cost
    if include_yield:
        col_widths = [30, 8, 12, 16, 30, 12, 8, 18, 18]
        headers = [
            "Crop Name", "Points", "Cycle Days", "Est. Yield (kg)",
            "Chemical", "Qty", "Unit", "Unit Price (AED)", "Total Cost (AED)",
        ]
        # Column indices (1-based)
        col_chem = 5
        col_qty = 6
        col_unit = 7
        col_unit_price = 8
        col_total_cost = 9
        num_cols = 9
    else:
        col_widths = [30, 8, 12, 30, 12, 8, 18, 18]
        headers = [
            "Crop Name", "Points", "Cycle Days",
            "Chemical", "Qty", "Unit", "Unit Price (AED)", "Total Cost (AED)",
        ]
        col_chem = 4
        col_qty = 5
        col_unit = 6
        col_unit_price = 7
        col_total_cost = 8
        num_cols = 8

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Header row ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_row = 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    grand_total = 0.0
    grand_total_known = True
    total_yield_by_unit: dict = {}  # yieldUnit → cumulative est yield

    for crop in response.perCrop:
        # -- Compute estimated yield for this crop if data available --
        est_yield = None
        yield_unit = None
        if include_yield:
            yi = yield_info_by_plant.get(str(crop.plantDataId))
            if yi:
                ypp = yi.get("yieldPerPlant", 0) or 0
                spp = yi.get("seedsPerPlantingPoint", 1) or 1
                waste = yi.get("expectedWastePercentage", 0) or 0
                per_dripper = ypp * spp * (1 - waste / 100)
                if per_dripper > 0:
                    est_yield = round(crop.points * per_dripper, 4)
                    yield_unit = yi.get("yieldUnit", "kg")
                    total_yield_by_unit[yield_unit] = total_yield_by_unit.get(yield_unit, 0) + est_yield

        # -- Crop header row --
        crop_fill = PatternFill("solid", fgColor=_CROP_BG)
        crop_font = Font(bold=True)
        crop_cells = [crop.plantName, crop.points, crop.cycleDays]
        for c_idx, val in enumerate(crop_cells, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.fill = crop_fill
            cell.font = crop_font
            # Apply integer thousands format to Points and Cycle Days
            if c_idx in (2, 3) and isinstance(val, (int, float)):
                cell.number_format = _FMT_INT
        if include_yield:
            yield_cell = ws.cell(row=current_row, column=4, value=est_yield if est_yield is not None else "—")
            yield_cell.fill = crop_fill
            yield_cell.font = crop_font
            if isinstance(est_yield, (int, float)):
                yield_cell.number_format = _FMT_DECIMAL
        # Remaining cols grey
        for c_idx in range(col_chem, num_cols + 1):
            ws.cell(row=current_row, column=c_idx).fill = crop_fill
        current_row += 1

        # -- Ingredient rows --
        for ing in crop.ingredients:
            ws.cell(row=current_row, column=col_chem, value=ing.name)
            qty_cell = ws.cell(row=current_row, column=col_qty, value=round(ing.qty, 4))
            qty_cell.number_format = _FMT_DECIMAL
            ws.cell(row=current_row, column=col_unit, value=ing.unit)
            up_cell = ws.cell(row=current_row, column=col_unit_price, value=ing.unitPrice)
            if isinstance(ing.unitPrice, (int, float)):
                up_cell.number_format = _FMT_DECIMAL
            tc_cell = ws.cell(row=current_row, column=col_total_cost, value=ing.totalCost)
            if isinstance(ing.totalCost, (int, float)):
                tc_cell.number_format = _FMT_DECIMAL
            current_row += 1

        # -- Subtotal row --
        total_fill = PatternFill("solid", fgColor=_TOTAL_BG)
        total_font = Font(bold=True)
        for c_idx in range(1, num_cols + 1):
            ws.cell(row=current_row, column=c_idx).fill = total_fill
        ws.cell(row=current_row, column=col_chem, value="Subtotal").font = total_font
        if crop.subtotalCost is not None:
            sub_cell = ws.cell(row=current_row, column=col_total_cost, value=round(crop.subtotalCost, 4))
            sub_cell.font = total_font
            sub_cell.number_format = _FMT_DECIMAL
        else:
            ws.cell(row=current_row, column=col_total_cost, value="N/A (missing prices)").font = total_font
            grand_total_known = False
        current_row += 1

        if crop.subtotalCost is not None and grand_total_known:
            grand_total += crop.subtotalCost

    # -- Total yield row (only when yield data was available) --
    if include_yield and total_yield_by_unit:
        yield_total_fill = PatternFill("solid", fgColor=_GRAND_BG)
        yield_total_font = Font(bold=True, size=12)
        for c_idx in range(1, num_cols + 1):
            ws.cell(row=current_row, column=c_idx).fill = yield_total_fill
        ws.cell(row=current_row, column=1, value="TOTAL YIELD").font = yield_total_font
        # Show first/only yieldUnit total in the Est. Yield column; if multiple
        # units somehow appear, format as "X kg + Y lbs" string in the same cell.
        if len(total_yield_by_unit) == 1:
            unit, total = next(iter(total_yield_by_unit.items()))
            cell = ws.cell(row=current_row, column=4, value=round(total, 4))
            cell.font = yield_total_font
            cell.number_format = _FMT_DECIMAL
            # Add unit suffix beside it in next col if available
            ws.cell(row=current_row, column=5, value=unit).font = yield_total_font
        else:
            label = " + ".join(f"{round(v, 4)} {u}" for u, v in total_yield_by_unit.items())
            cell = ws.cell(row=current_row, column=4, value=label)
            cell.font = yield_total_font
        current_row += 1

    # -- Grand total cost row --
    gt_fill = PatternFill("solid", fgColor=_GRAND_BG)
    gt_font = Font(bold=True, size=12)
    for c_idx in range(1, num_cols + 1):
        ws.cell(row=current_row, column=c_idx).fill = gt_fill
    ws.cell(row=current_row, column=1, value="GRAND TOTAL").font = gt_font
    if grand_total_known:
        gt_cell = ws.cell(row=current_row, column=col_total_cost, value=round(grand_total, 4))
        gt_cell.font = gt_font
        gt_cell.number_format = _FMT_DECIMAL
    else:
        ws.cell(row=current_row, column=col_total_cost, value="N/A (missing prices)").font = gt_font

    # -- Per Input sheet --
    _write_per_input_sheet(wb, response)

    # -- Warnings sheet --
    if response.warnings:
        ws_warn = wb.create_sheet(title="Warnings")
        ws_warn.column_dimensions["A"].width = 80
        ws_warn.cell(row=1, column=1, value="Warnings").font = Font(bold=True)
        for i, w in enumerate(response.warnings, start=2):
            ws_warn.cell(row=i, column=1, value=w)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_per_input_sheet(wb: Workbook, response: CalculateResponse) -> None:
    """
    Add a "Per Input" sheet that aggregates ingredient quantities and costs
    across all crops in the calculation response.

    Aggregation key: chemicalId when present, else (name, unit) for unmatched
    or archived-chemical rows.
    """
    ws = wb.create_sheet(title="Per Input")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    headers = ["Chemical", "Total Qty", "Unit", "Unit Price (AED)", "Total Cost (AED)"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Aggregate: key → {name, qty, unit, unitPrice, totalCost, costKnown}
    aggregated: dict = {}
    for crop in response.perCrop:
        for ing in crop.ingredients:
            key = str(ing.chemicalId) if ing.chemicalId else f"unmatched::{ing.name}::{ing.unit}"
            entry = aggregated.setdefault(
                key,
                {
                    "name": ing.name,
                    "unit": ing.unit,
                    "qty": 0.0,
                    "unitPrice": ing.unitPrice,
                    "totalCost": 0.0,
                    "costKnown": True,
                },
            )
            entry["qty"] += ing.qty
            if ing.totalCost is not None:
                entry["totalCost"] += ing.totalCost
            else:
                entry["costKnown"] = False
            # Reason: keep the last seen unitPrice; all rows for the same chemical
            # share the same chemical.defaultUnit price.
            if ing.unitPrice is not None:
                entry["unitPrice"] = ing.unitPrice

    current_row = 2
    grand_total = 0.0
    grand_total_known = True
    for entry in sorted(aggregated.values(), key=lambda e: e["name"].lower()):
        ws.cell(row=current_row, column=1, value=entry["name"])
        qty_cell = ws.cell(row=current_row, column=2, value=round(entry["qty"], 4))
        qty_cell.number_format = _FMT_DECIMAL
        ws.cell(row=current_row, column=3, value=entry["unit"])
        up_cell = ws.cell(row=current_row, column=4, value=entry["unitPrice"])
        if isinstance(entry["unitPrice"], (int, float)):
            up_cell.number_format = _FMT_DECIMAL
        if entry["costKnown"] and entry["unitPrice"] is not None:
            tc_cell = ws.cell(row=current_row, column=5, value=round(entry["totalCost"], 4))
            tc_cell.number_format = _FMT_DECIMAL
            grand_total += entry["totalCost"]
        else:
            ws.cell(row=current_row, column=5, value="N/A")
            grand_total_known = False
        current_row += 1

    # Grand total row
    gt_fill = PatternFill("solid", fgColor=_GRAND_BG)
    gt_font = Font(bold=True, size=12)
    for c_idx in range(1, 6):
        ws.cell(row=current_row, column=c_idx).fill = gt_fill
    ws.cell(row=current_row, column=1, value="GRAND TOTAL").font = gt_font
    if grand_total_known:
        gt_cell = ws.cell(row=current_row, column=5, value=round(grand_total, 4))
        gt_cell.font = gt_font
        gt_cell.number_format = _FMT_DECIMAL
    else:
        ws.cell(row=current_row, column=5, value="N/A (missing prices)").font = gt_font


def _is_net_yield_header(header_str: str) -> bool:
    """
    Return True if header_str matches the "Net Yield" family of column names.

    Accepted (case-insensitive):
      "Net Yield", "Net Yield (kg)", "net yield", "NET YIELD (KG)", ...

    Args:
        header_str: Stripped string value of a header cell.

    Returns:
        True if the string is a recognised Net Yield column label.
    """
    return bool(_NET_YIELD_PATTERN.match(header_str.strip()))


def _try_parse_positive_float(value: object) -> Optional[float]:
    """
    Attempt to parse *value* as a positive (> 0) float.

    Returns the float on success, or None if the value is absent, zero,
    negative, or unparseable.

    Args:
        value: Raw cell value from the worksheet.

    Returns:
        Positive float, or None.
    """
    if value is None:
        return None
    try:
        fval = float(value)
    except (TypeError, ValueError):
        return None
    return fval if fval > 0 else None


async def import_crops(file_bytes: bytes) -> ParsedImport:
    """
    Parse a .xlsx file to extract crop name + points pairs.

    Expected format:
    - First sheet
    - First row: headers — must contain at least "Crop Name" and "Points"
      (case-insensitive, leading/trailing whitespace ignored).
      Optionally "Net Yield (kg)" (case-insensitive, trailing unit in parens optional).
    - Subsequent rows: data

    Processing rules per data row:
    1. Resolve crop name to a plant_data_enhanced document via case-insensitive match.
       Skip with reason "Unknown crop" if not found.
    2. If Net Yield (C) column exists and the cell is a positive number:
       a. Compute yieldPerDripper = yieldPerPlant × seedsPerPlantingPoint
                                    × (1 − expectedWastePercentage / 100).
       b. If yieldPerDripper <= 0, skip with reason "Plant has invalid yield rate".
       c. Compute points = ceil(netYield / yieldPerDripper).
       d. If plant's yieldUnit is not "kg", append an informational warning.
       e. Ignore the Points column for this row.
    3. Else if Net Yield (C) is present in the column but the cell value is not a
       positive number AND not None:
       a. If the cell is a non-numeric string → skip with "Net Yield is not a number".
       b. If the cell is 0 / negative → fall through to check Points.
    4. Else (no Net Yield column, or cell is empty/zero): check Points.
    5. If Points (B) is set and positive: use as-is.
    6. Else: skip with reason "Row has neither Points nor Net Yield" (or specific parse
       error if the value is non-numeric).
    7. If computed/provided points exceed 10,000,000: clamp to 10,000,000 and add a warning.

    Aggregate duplicate plantDataIds by summing resulting points (unchanged from v1 behaviour).

    Args:
        file_bytes: Raw .xlsx file bytes.

    Returns:
        ParsedImport with items, skipped, and warnings.

    Raises:
        ValueError: If the file cannot be parsed or has no recognisable header row.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file: {exc}") from exc

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty — no rows found")

    # -- Find and parse header row --
    header_row = rows[0]
    header_lower = [
        str(cell).strip().lower() if cell is not None else ""
        for cell in header_row
    ]

    if "crop name" not in header_lower or "points" not in header_lower:
        raise ValueError(
            "Header row must contain 'Crop Name' and 'Points' columns "
            f"(found: {header_lower})"
        )

    crop_col = header_lower.index("crop name")
    points_col = header_lower.index("points")

    # Reason: Net Yield column is optional — old-format files (2-col) still work.
    net_yield_col: Optional[int] = None
    for idx, h in enumerate(header_row):
        raw_h = str(h).strip() if h is not None else ""
        if _is_net_yield_header(raw_h):
            net_yield_col = idx
            break

    db = farm_db.get_database()

    items_by_plant: dict = {}  # plantDataId → ParsedImportItem
    skipped: List[SkippedRow] = []
    import_warnings: List[str] = []

    for row_idx, row in enumerate(rows[1:], start=1):
        raw_name = row[crop_col] if crop_col < len(row) else None
        raw_points = row[points_col] if points_col < len(row) else None
        raw_net_yield = (
            row[net_yield_col]
            if net_yield_col is not None and net_yield_col < len(row)
            else None
        )

        # Skip completely blank rows silently
        if raw_name is None and raw_points is None and raw_net_yield is None:
            continue

        name_str = str(raw_name).strip() if raw_name is not None else ""
        if not name_str:
            skipped.append(SkippedRow(rowIndex=row_idx, name="", reason="Empty crop name"))
            continue

        # -- Resolve plant --
        plant_doc = await db.plant_data_enhanced.find_one(
            {
                "deletedAt": None,
                "plantName": {
                    "$regex": f"^{_escape_regex_excel(name_str)}$",
                    "$options": "i",
                },
            },
            {
                "plantDataId": 1,
                "plantName": 1,
                "yieldInfo": 1,
            },
        )

        if plant_doc is None:
            skipped.append(
                SkippedRow(rowIndex=row_idx, name=name_str, reason="Unknown crop")
            )
            continue

        pid = plant_doc["plantDataId"]
        plant_name = plant_doc["plantName"]

        # -- Determine points via Net Yield or Points column --
        points: Optional[int] = None

        if net_yield_col is not None:
            # Net Yield column exists — evaluate the cell
            if raw_net_yield is not None:
                net_yield_positive = _try_parse_positive_float(raw_net_yield)

                if net_yield_positive is None:
                    # Check whether it is a non-numeric string (parse error) or
                    # a non-positive number (fall through to Points)
                    try:
                        float(raw_net_yield)
                        # Parsed but <= 0 — treat as not set; fall through to Points
                    except (TypeError, ValueError):
                        # Cannot parse at all — skip the row
                        skipped.append(
                            SkippedRow(
                                rowIndex=row_idx,
                                name=name_str,
                                reason="Net Yield is not a number",
                            )
                        )
                        continue
                else:
                    # Positive Net Yield value — convert to points via yieldInfo
                    yield_info = plant_doc.get("yieldInfo") or {}
                    yield_per_plant = float(yield_info.get("yieldPerPlant", 0) or 0)
                    seeds_per_point = int(yield_info.get("seedsPerPlantingPoint", 1) or 1)
                    waste_pct = float(yield_info.get("expectedWastePercentage", 0) or 0)
                    yield_unit = str(yield_info.get("yieldUnit", "kg") or "kg")

                    # Reason: yieldPerDripper is the net harvestable yield per
                    # planting point after accounting for waste.
                    yield_per_dripper = (
                        yield_per_plant
                        * seeds_per_point
                        * (1.0 - waste_pct / 100.0)
                    )

                    if yield_per_dripper <= 0:
                        skipped.append(
                            SkippedRow(
                                rowIndex=row_idx,
                                name=name_str,
                                reason="Plant has invalid yield rate",
                            )
                        )
                        continue

                    # Warn if the unit is not kg (user input is assumed to match the unit)
                    if yield_unit.lower() != "kg":
                        import_warnings.append(
                            f"Crop '{plant_name}' has yieldUnit='{yield_unit}' "
                            f"— Net Yield input was interpreted as {yield_unit}"
                        )

                    computed_points = math.ceil(net_yield_positive / yield_per_dripper)
                    if computed_points > _MAX_POINTS:
                        import_warnings.append(
                            f"Row {row_idx} ('{plant_name}'): computed points "
                            f"{computed_points:,} exceeded maximum {_MAX_POINTS:,} "
                            f"— clamped to {_MAX_POINTS:,}"
                        )
                        computed_points = _MAX_POINTS

                    points = computed_points

        if points is None:
            # Fall through: use Points column
            if raw_points is None:
                skipped.append(
                    SkippedRow(
                        rowIndex=row_idx,
                        name=name_str,
                        reason="Row has neither Points nor Net Yield",
                    )
                )
                continue

            try:
                points_val = int(raw_points)
                if points_val < 1:
                    raise ValueError("Must be >= 1")
            except (TypeError, ValueError):
                try:
                    float(raw_points)
                    # It is a number but not a valid integer >= 1
                    skipped.append(
                        SkippedRow(
                            rowIndex=row_idx,
                            name=name_str,
                            reason=(
                                f"Invalid points value: '{raw_points}' "
                                f"(must be a positive integer)"
                            ),
                        )
                    )
                except (TypeError, ValueError):
                    skipped.append(
                        SkippedRow(
                            rowIndex=row_idx,
                            name=name_str,
                            reason="Points is not a number",
                        )
                    )
                continue

            if points_val > _MAX_POINTS:
                import_warnings.append(
                    f"Row {row_idx} ('{plant_name}'): points value "
                    f"{points_val:,} exceeded maximum {_MAX_POINTS:,} "
                    f"— clamped to {_MAX_POINTS:,}"
                )
                points_val = _MAX_POINTS

            points = points_val

        # -- Aggregate by plant --
        if pid in items_by_plant:
            items_by_plant[pid].points += points
        else:
            items_by_plant[pid] = ParsedImportItem(
                plantDataId=pid,
                plantName=plant_name,
                points=points,
            )

    if skipped:
        import_warnings.append(
            f"{len(skipped)} row(s) were skipped — see the 'skipped' list for details"
        )

    return ParsedImport(
        items=list(items_by_plant.values()),
        skipped=skipped,
        warnings=import_warnings,
    )


def _escape_regex_excel(text: str) -> str:
    """Escape special MongoDB regex characters."""
    special = r"\.^$*+?{}[]|()"
    return "".join(f"\\{c}" if c in special else c for c in text)
