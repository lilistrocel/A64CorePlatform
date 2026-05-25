"""
Finance Report Export API (T-060.6)

Provides streaming download endpoints for the three statutory financial
statements — Balance Sheet, Income Statement, and Cash Flow Statement.

Endpoint:
  GET /reports/export/{statement}?format=pdf|xlsx

Path parameters:
  statement: balance-sheet | income-statement | cash-flow

Query parameters:
  format:                 pdf | xlsx  (required)
  organization_id:        required
  company_code:           required
  --- Balance Sheet ---
  as_of_date:             optional date (default: today)
  include_voided:         bool (default: false)
  cost_center_id:         optional str
  --- Income Statement / Cash Flow ---
  period_start:           required date
  period_end:             required date
  --- Income Statement only ---
  compare_period_start:   optional date
  compare_period_end:     optional date
  --- Shared ---
  cost_center_id:         optional str (repeatable — ?cost_center_id=A&cost_center_id=B)
  include_voided:         bool (default: false)

Permissions:
  Same read roles as the JSON report endpoints
  (accountant, finance_admin, auditor, admin, super_admin).

Security:
  - No user-supplied data is executed as code.
  - File content is generated server-side from DB data only.
  - Content-Disposition attachment header prevents inline execution.
  - WeasyPrint renders server-side HTML — no user HTML accepted.
"""

import io
import logging
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from ...db.session import get_db
from ...middleware.auth import TokenPayload, require_roles
from .reports import (
    _READ_ROLES,
    get_balance_sheet,
    get_cash_flow,
    get_income_statement,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Reports"])

# ---------------------------------------------------------------------------
# Allowed path-param values
# ---------------------------------------------------------------------------

_VALID_STATEMENTS = frozenset({"balance-sheet", "income-statement", "cash-flow"})
_VALID_FORMATS = frozenset({"pdf", "xlsx"})

# ---------------------------------------------------------------------------
# Template directory (sibling to this file)
# ---------------------------------------------------------------------------

_TEMPLATE_DIR = Path(__file__).parent / "templates"

# ---------------------------------------------------------------------------
# Amount formatting helpers
# ---------------------------------------------------------------------------

_PAREN_RE = re.compile(r"^-(.+)$")


def _fmt_amount_str(value: str) -> str:
    """
    Format a Decimal string as a human-readable amount with thousands
    separator.  Negative values are rendered in parentheses per
    accounting convention — e.g. '-1234.56' → '(1,234.56)'.

    Args:
        value: Decimal string (as produced by the report endpoints).

    Returns:
        Formatted string, e.g. '50,000.00' or '(1,234.56)'.
    """
    try:
        d = Decimal(str(value))
    except InvalidOperation:
        return str(value)

    abs_d = abs(d)
    # Format with 2 decimal places and thousands separator.
    formatted = f"{abs_d:,.2f}"
    if d < Decimal("0"):
        return f"({formatted})"
    return formatted


def _fmt_amount_decimal(value: Decimal) -> str:
    """Wrapper for Decimal inputs (used in Excel builders)."""
    return _fmt_amount_str(str(value))


# ---------------------------------------------------------------------------
# Safe filename component builder
# ---------------------------------------------------------------------------

def _safe_slug(text: str) -> str:
    """
    Replace any character that is not alphanumeric, hyphen, or underscore
    with an underscore.  Limits the result to 40 characters.

    Args:
        text: Raw string (e.g. company code, org name, period).

    Returns:
        Safe slug suitable for use in a filename.
    """
    return re.sub(r"[^\w\-]", "_", text)[:40]


def _build_filename(statement: str, period_label: str, company_code: str, ext: str) -> str:
    """
    Build a download filename.

    Convention: {statement}_{period_label}_{company_code}.{ext}

    Args:
        statement: e.g. 'balance-sheet'
        period_label: e.g. '2026-12-31' or '2026-01-01_2026-12-31'
        company_code: company code string
        ext: 'pdf' or 'xlsx'

    Returns:
        Filename string, e.g. 'balance-sheet_2026-12-31_A001.pdf'
    """
    slug_statement = _safe_slug(statement)
    slug_period = _safe_slug(period_label)
    slug_company = _safe_slug(company_code)
    return f"{slug_statement}_{slug_period}_{slug_company}.{ext}"


# ---------------------------------------------------------------------------
# Excel (openpyxl) builders — one per statement
# ---------------------------------------------------------------------------

def _excel_header_font(wb):
    """Return an openpyxl Font configured for bold section headers."""
    from openpyxl.styles import Font
    return Font(bold=True, color="1A3B6E")


def _apply_header_fill(cell, wb):
    """Apply a light-blue fill to header cells."""
    from openpyxl.styles import PatternFill
    cell.fill = PatternFill(
        start_color="DCE6F4", end_color="DCE6F4", fill_type="solid"
    )


def _apply_total_fill(cell):
    """Apply a slightly darker fill to total/subtotal cells."""
    from openpyxl.styles import PatternFill
    cell.fill = PatternFill(
        start_color="F0F4F9", end_color="F0F4F9", fill_type="solid"
    )


def _write_tenant_header(ws, org_name: str, company_code: str,
                          title: str, subtitle: str, generated_at: str) -> int:
    """
    Write a 5-row letterhead block at the top of the worksheet.

    Returns the next available row number.

    Args:
        ws: openpyxl Worksheet.
        org_name: Organisation / tenant name.
        company_code: Company code string.
        title: Statement title.
        subtitle: Period/date subtitle.
        generated_at: ISO datetime string.

    Returns:
        Row index immediately after the header block.
    """
    from openpyxl.styles import Font, Alignment

    ws.cell(row=1, column=1, value=org_name).font = Font(
        bold=True, size=14, color="1A3B6E"
    )
    ws.cell(row=2, column=1, value=f"Company: {company_code}").font = Font(size=10)
    ws.cell(row=3, column=1, value=title).font = Font(
        bold=True, size=12, color="1A3B6E"
    )
    ws.cell(row=4, column=1, value=subtitle).font = Font(italic=True, size=10)
    ws.cell(row=5, column=1, value=f"Generated: {generated_at}").font = Font(
        size=8, color="888888"
    )
    # Blank separator row.
    return 7


def _section_row(ws, row_num: int, label: str, col_count: int) -> None:
    """
    Write a section header row spanning all columns (bold, dark blue bg).

    Args:
        ws: Worksheet.
        row_num: Target row.
        label: Section label text.
        col_count: Number of data columns (determines merge width).
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row_num, column=1, value=label.upper())
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill(start_color="1A3B6E", end_color="1A3B6E", fill_type="solid")
    ws.merge_cells(
        start_row=row_num, start_column=1,
        end_row=row_num, end_column=1 + col_count,
    )


def _amount_cell(ws, row_num: int, col: int, value_str: str,
                 bold: bool = False, fill: bool = False) -> None:
    """
    Write a formatted amount into a cell.

    Args:
        ws: Worksheet.
        row_num: Target row.
        col: Target column (1-based).
        value_str: Decimal string (may be negative).
        bold: Whether to apply bold font.
        fill: Whether to apply the total-row fill.
    """
    from openpyxl.styles import Font, Alignment
    try:
        raw = Decimal(str(value_str))
        cell = ws.cell(row=row_num, column=col, value=float(raw))
        # Reason: Excel number format with parentheses for negatives,
        # thousands separator, 2 decimal places.
        cell.number_format = '#,##0.00_);(#,##0.00)'
    except InvalidOperation:
        cell = ws.cell(row=row_num, column=col, value=value_str)

    cell.alignment = __import__("openpyxl").styles.Alignment(horizontal="right")
    if bold:
        cell.font = Font(bold=True)
    if fill:
        _apply_total_fill(cell)


def _build_balance_sheet_xlsx(data: dict, company_code: str) -> bytes:
    """
    Build an openpyxl workbook for the Balance Sheet.

    Args:
        data: The 'data' dict from the BalanceSheetResponse JSON.
        company_code: Company code string.

    Returns:
        Raw bytes of the .xlsx file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # Column widths.
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    subtitle = f"As of {data['asOfDate']}"
    r = _write_tenant_header(
        ws,
        org_name=data.get("organizationId", ""),
        company_code=company_code,
        title="Balance Sheet",
        subtitle=subtitle,
        generated_at=data["generatedAt"],
    )

    # Column headers.
    for col, label in [(1, "Account"), (2, "Balance (AED)"), (3, "Account No.")]:
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A3B6E", end_color="1A3B6E", fill_type="solid")
    r += 1

    rows = data.get("rows", [])
    totals = data.get("totals", {})

    for drawer_label, drawer_key, total_key in [
        ("ASSETS", "ASSETS", "totalAssets"),
        ("LIABILITIES", "LIABILITIES", "totalLiabilities"),
        ("EQUITY", "EQUITY", "totalEquity"),
    ]:
        _section_row(ws, r, drawer_label, 2)
        r += 1

        for row in rows:
            if row["drawer"] != drawer_key:
                continue
            indent = "    " if not row["isHeader"] else ""
            label_cell = ws.cell(row=r, column=1, value=f"{indent}{row['accountName']}")
            if row["isHeader"]:
                label_cell.font = Font(bold=True, color="1A3B6E")
                _apply_header_fill(label_cell, wb)
            _amount_cell(ws, r, 2, row["balance"], bold=row["isHeader"], fill=row["isHeader"])
            ws.cell(row=r, column=3, value=row["accountNumber"])
            r += 1

        # Synthetic current-year P/L row inside equity section.
        if drawer_key == "EQUITY":
            ws.cell(row=r, column=1, value="    Current Year Profit / (Loss)").font = Font(
                italic=True
            )
            _amount_cell(ws, r, 2, data.get("currentYearProfitLoss", "0"))
            r += 1

        # Subtotal row.
        total_val = totals.get(total_key, "0")
        subtotal_cell = ws.cell(row=r, column=1, value=f"Total {drawer_label.title()}")
        subtotal_cell.font = Font(bold=True)
        _apply_total_fill(subtotal_cell)
        _amount_cell(ws, r, 2, total_val, bold=True, fill=True)
        _apply_total_fill(ws.cell(row=r, column=3))
        r += 1

        # Blank separator.
        r += 1

    # Grand total — Total Liabilities + Equity.
    grand_cell = ws.cell(row=r, column=1, value="Total Liabilities + Equity")
    grand_cell.font = Font(bold=True, size=10, color="1A3B6E")
    _apply_header_fill(grand_cell, wb)
    _amount_cell(ws, r, 2, totals.get("totalLiabilitiesPlusEquity", "0"), bold=True, fill=True)
    r += 2

    # Warnings.
    for warning in data.get("warnings", []):
        ws.cell(row=r, column=1, value=f"WARNING: {warning}").font = Font(
            color="CC5500", italic=True
        )
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_income_statement_xlsx(data: dict, company_code: str) -> bytes:
    """
    Build an openpyxl workbook for the Income Statement.

    Args:
        data: The 'data' dict from the IncomeStatementResponse JSON.
        company_code: Company code string.

    Returns:
        Raw bytes of the .xlsx file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    primary = data["primary"]
    comparison = data.get("comparison")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    if comparison:
        ws.column_dimensions["C"].width = 20

    subtitle = f"{primary['periodStart']} to {primary['periodEnd']}"
    if comparison:
        subtitle += f"  |  vs  {comparison['periodStart']} to {comparison['periodEnd']}"

    r = _write_tenant_header(
        ws,
        org_name=data.get("organizationId", ""),
        company_code=company_code,
        title="Income Statement",
        subtitle=subtitle,
        generated_at=data["generatedAt"],
    )

    # Column headers.
    headers = [
        (1, "Description"),
        (2, f"{primary['periodStart']} – {primary['periodEnd']}"),
    ]
    if comparison:
        headers.append((3, f"{comparison['periodStart']} – {comparison['periodEnd']}"))

    for col, label in headers:
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A3B6E", end_color="1A3B6E", fill_type="solid")
    r += 1

    _DRAWER_LABELS = {
        "REVENUE": "Revenue",
        "COST_OF_SALES": "Cost of Sales",
        "OPERATING_COST": "Operating Expenses",
        "OTHER_INCOME": "Other Income",
        "NON_OPERATING": "Non-Operating Items",
        "TAXATION": "Taxation",
    }

    comp_sections_by_drawer = {}
    if comparison:
        for s in comparison["sections"]:
            comp_sections_by_drawer[s["drawer"]] = s

    subtotals = primary["subtotals"]
    comp_subtotals = comparison["subtotals"] if comparison else None

    for section in primary["sections"]:
        drawer = section["drawer"]
        label = _DRAWER_LABELS.get(drawer, drawer)

        _section_row(ws, r, label, 2 if not comparison else 3)
        r += 1

        comp_section = comp_sections_by_drawer.get(drawer)
        comp_rows_by_id = {}
        if comp_section:
            for cr in comp_section["rows"]:
                comp_rows_by_id[cr["accountId"]] = cr

        for row in section["rows"]:
            if row["isHeader"]:
                continue
            label_cell = ws.cell(row=r, column=1, value=f"    {row['accountName']}")
            _amount_cell(ws, r, 2, row["balance"])
            if comparison:
                comp_row = comp_rows_by_id.get(row["accountId"])
                _amount_cell(ws, r, 3, comp_row["balance"] if comp_row else "0")
            r += 1

        # Section total.
        subtotal_label = ws.cell(row=r, column=1, value=f"Total {label}")
        subtotal_label.font = Font(bold=True)
        _apply_total_fill(subtotal_label)
        _amount_cell(ws, r, 2, section["total"], bold=True, fill=True)
        if comparison and comp_section:
            _amount_cell(ws, r, 3, comp_section["total"], bold=True, fill=True)
        r += 1

        # Key subtotal rows.
        if drawer == "COST_OF_SALES":
            gp_cell = ws.cell(row=r, column=1, value="Gross Profit")
            gp_cell.font = Font(bold=True, color="1A3B6E")
            _apply_header_fill(gp_cell, wb)
            _amount_cell(ws, r, 2, subtotals["grossProfit"], bold=True, fill=True)
            if comparison and comp_subtotals:
                _amount_cell(ws, r, 3, comp_subtotals["grossProfit"], bold=True, fill=True)
            r += 1

        if drawer == "OPERATING_COST":
            ebit_cell = ws.cell(row=r, column=1, value="Operating Income (EBIT)")
            ebit_cell.font = Font(bold=True, color="1A3B6E")
            _apply_header_fill(ebit_cell, wb)
            _amount_cell(ws, r, 2, subtotals["operatingIncome"], bold=True, fill=True)
            if comparison and comp_subtotals:
                _amount_cell(ws, r, 3, comp_subtotals["operatingIncome"], bold=True, fill=True)
            r += 1

        r += 1  # blank separator

    # Net Income.
    ni_cell = ws.cell(row=r, column=1, value="Net Income / (Loss)")
    ni_cell.font = Font(bold=True, size=11, color="1A3B6E")
    _apply_header_fill(ni_cell, wb)
    _amount_cell(ws, r, 2, subtotals["netIncome"], bold=True, fill=True)
    if comparison and comp_subtotals:
        _amount_cell(ws, r, 3, comp_subtotals["netIncome"], bold=True, fill=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_cash_flow_xlsx(data: dict, company_code: str) -> bytes:
    """
    Build an openpyxl workbook for the Cash Flow Statement.

    Args:
        data: The 'data' dict from the CashFlowResponse JSON.
        company_code: Company code string.

    Returns:
        Raw bytes of the .xlsx file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 20

    subtitle = f"{data['periodStart']} to {data['periodEnd']}"
    r = _write_tenant_header(
        ws,
        org_name=data.get("organizationId", ""),
        company_code=company_code,
        title="Cash Flow Statement (Indirect Method)",
        subtitle=subtitle,
        generated_at=data["generatedAt"],
    )

    # Column headers.
    for col, label in [(1, "Description"), (2, "Amount (AED)")]:
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A3B6E", end_color="1A3B6E", fill_type="solid")
    r += 1

    operating = data["operating"]
    investing = data["investing"]
    financing = data["financing"]

    def _section(label: str) -> None:
        nonlocal r
        _section_row(ws, r, label, 1)
        r += 1

    def _line(label: str, value: str, indent: int = 1,
              bold: bool = False, fill: bool = False) -> None:
        nonlocal r
        prefix = "    " * indent
        c = ws.cell(row=r, column=1, value=f"{prefix}{label}")
        if bold:
            c.font = Font(bold=True)
        if fill:
            _apply_total_fill(c)
        _amount_cell(ws, r, 2, value, bold=bold, fill=fill)
        r += 1

    # ── Operating ──────────────────────────────────────────────────────
    _section("Operating Activities")
    _line("Net Income / (Loss)", operating["netIncome"])

    if operating["nonCashAdjustments"]:
        ws.cell(row=r, column=1, value="    Adjustments for Non-Cash Items").font = Font(bold=True)
        r += 1
        for line in operating["nonCashAdjustments"]:
            _line(line["accountName"], line["contribution"], indent=2)
        _line("Total Non-Cash Adjustments", operating["nonCashAdjustmentsTotal"],
              bold=True, fill=True)

    if operating["workingCapitalChanges"]:
        ws.cell(row=r, column=1, value="    Changes in Working Capital").font = Font(bold=True)
        r += 1
        for line in operating["workingCapitalChanges"]:
            _line(line["accountName"], line["contribution"], indent=2)
        _line("Total Working Capital Changes", operating["workingCapitalChangesTotal"],
              bold=True, fill=True)

    _line("Net Cash from Operating Activities", operating["total"], indent=0, bold=True, fill=True)
    r += 1  # blank

    # ── Investing ──────────────────────────────────────────────────────
    _section("Investing Activities")
    if investing["items"]:
        for line in investing["items"]:
            _line(line["accountName"], line["contribution"])
    else:
        ws.cell(row=r, column=1, value="    No investing activity").font = Font(italic=True)
        r += 1
    _line("Net Cash from Investing Activities", investing["total"], indent=0, bold=True, fill=True)
    r += 1  # blank

    # ── Financing ──────────────────────────────────────────────────────
    _section("Financing Activities")
    if financing["items"]:
        for line in financing["items"]:
            _line(line["accountName"], line["contribution"])
    else:
        ws.cell(row=r, column=1, value="    No financing activity").font = Font(italic=True)
        r += 1
    _line("Net Cash from Financing Activities", financing["total"], indent=0, bold=True, fill=True)
    r += 1  # blank

    # ── Cash position ─────────────────────────────────────────────────
    _section("Cash Position")
    c = ws.cell(row=r, column=1, value="Net Change in Cash")
    c.font = Font(bold=True, size=10, color="1A3B6E")
    _apply_header_fill(c, wb)
    _amount_cell(ws, r, 2, data["netChangeInCash"], bold=True, fill=True)
    r += 1
    _line("Cash at Beginning of Period", data["cashAtBeginning"])
    _line("Cash at End of Period", data["cashAtEnd"], bold=True, fill=True)
    r += 1

    for warning in data.get("warnings", []):
        ws.cell(row=r, column=1, value=f"WARNING: {warning}").font = Font(
            color="CC5500", italic=True
        )
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF (WeasyPrint + Jinja2) builders
# ---------------------------------------------------------------------------

def _jinja_env():
    """
    Return (and cache) a Jinja2 Environment configured to load templates
    from the `templates/` directory next to this module.

    The `fmt_amount` filter is registered so templates can call
    {{ value | fmt_amount }}.

    Returns:
        jinja2.Environment instance.
    """
    from jinja2 import Environment, FileSystemLoader, select_autoescape
    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATE_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    env.filters["fmt_amount"] = _fmt_amount_str
    return env


def _render_pdf(template_name: str, context: dict) -> bytes:
    """
    Render a Jinja2 HTML template to PDF bytes via WeasyPrint.

    Args:
        template_name: Filename inside the templates/ directory.
        context: Template context variables.

    Returns:
        PDF bytes.

    Raises:
        RuntimeError: If WeasyPrint or Jinja2 fails.
    """
    from weasyprint import HTML

    env = _jinja_env()
    tmpl = env.get_template(template_name)
    html_str = tmpl.render(**context)
    # Reason: use base_url=str(_TEMPLATE_DIR) so relative CSS/font URLs
    # inside the template resolve correctly when WeasyPrint fetches them.
    pdf_bytes: bytes = HTML(
        string=html_str,
        base_url=str(_TEMPLATE_DIR),
    ).write_pdf()
    return pdf_bytes


def _build_balance_sheet_pdf(data: dict) -> bytes:
    """
    Render the Balance Sheet PDF.

    Args:
        data: The 'data' dict from the BalanceSheetResponse JSON.

    Returns:
        PDF bytes.
    """
    context = {
        "title": "Balance Sheet",
        "subtitle": f"As of {data['asOfDate']}",
        "org_name": data.get("organizationId", ""),
        "company_code": data.get("companyCode", ""),
        "currency": data.get("currency", "AED"),
        "generated_at": data.get("generatedAt", ""),
        "includes_voided": data.get("includesVoided", False),
        "cost_center_label": None,  # populated by caller if needed
        "rows": data.get("rows", []),
        "current_year_pl": data.get("currentYearProfitLoss", "0"),
        "totals": data.get("totals", {}),
        "warnings": data.get("warnings", []),
    }
    return _render_pdf("balance_sheet.html", context)


def _build_income_statement_pdf(data: dict) -> bytes:
    """
    Render the Income Statement PDF.

    Args:
        data: The 'data' dict from the IncomeStatementResponse JSON.

    Returns:
        PDF bytes.
    """
    primary = data["primary"]
    comparison = data.get("comparison")
    subtitle = f"{primary['periodStart']} to {primary['periodEnd']}"
    if comparison:
        subtitle += f"  |  Comparative: {comparison['periodStart']} to {comparison['periodEnd']}"

    context = {
        "title": "Income Statement",
        "subtitle": subtitle,
        "org_name": data.get("organizationId", ""),
        "company_code": data.get("companyCode", ""),
        "currency": data.get("currency", "AED"),
        "generated_at": data.get("generatedAt", ""),
        "includes_voided": data.get("includesVoided", False),
        "cost_center_label": None,
        "primary": primary,
        "comparison": comparison,
        "warnings": data.get("warnings", []),
    }
    return _render_pdf("income_statement.html", context)


def _build_cash_flow_pdf(data: dict) -> bytes:
    """
    Render the Cash Flow Statement PDF.

    Args:
        data: The 'data' dict from the CashFlowResponse JSON.

    Returns:
        PDF bytes.
    """
    context = {
        "title": "Cash Flow Statement (Indirect Method)",
        "subtitle": f"{data['periodStart']} to {data['periodEnd']}",
        "org_name": data.get("organizationId", ""),
        "company_code": data.get("companyCode", ""),
        "currency": data.get("currency", "AED"),
        "generated_at": data.get("generatedAt", ""),
        "includes_voided": data.get("includesVoided", False),
        "cost_center_label": None,
        "operating": data.get("operating", {}),
        "investing": data.get("investing", {}),
        "financing": data.get("financing", {}),
        "net_change_in_cash": data.get("netChangeInCash", "0"),
        "cash_at_beginning": data.get("cashAtBeginning", "0"),
        "cash_at_end": data.get("cashAtEnd", "0"),
        "warnings": data.get("warnings", []),
    }
    return _render_pdf("cash_flow.html", context)


# ---------------------------------------------------------------------------
# Export endpoint
# ---------------------------------------------------------------------------

@router.get(
    "/reports/export/{statement}",
    summary="Export financial statement (PDF or Excel)",
    description=(
        "Wave 2 (T-060.6) — Streaming download for the three statutory "
        "financial statements.\n\n"
        "**statement** path param: `balance-sheet` | `income-statement` | `cash-flow`\n\n"
        "**format** query param: `pdf` | `xlsx`\n\n"
        "All other query parameters mirror the corresponding JSON report "
        "endpoint exactly — the same data is used to produce both the JSON "
        "view and the exported file, guaranteeing they match.\n\n"
        "Returns a streaming `application/pdf` or "
        "`application/vnd.openxmlformats-officedocument.spreadsheetml.sheet` "
        "response with `Content-Disposition: attachment`."
    ),
    response_class=StreamingResponse,
)
async def export_report(
    statement: str,
    # ── Required params ────────────────────────────────────────────────
    format: str = Query(..., description="Output format: pdf | xlsx"),
    organization_id: str = Query(..., description="Organisation scope"),
    company_code: str = Query(..., description="Company code"),
    # ── Balance Sheet ──────────────────────────────────────────────────
    as_of_date: Optional[date] = Query(
        None,
        description="(Balance Sheet) Snapshot date. Defaults to today.",
    ),
    # ── Income Statement + Cash Flow ───────────────────────────────────
    period_start: Optional[date] = Query(
        None,
        description="(Income Statement / Cash Flow) Period start (inclusive).",
    ),
    period_end: Optional[date] = Query(
        None,
        description="(Income Statement / Cash Flow) Period end (inclusive).",
    ),
    # ── Income Statement only ──────────────────────────────────────────
    compare_period_start: Optional[date] = Query(
        None,
        description="(Income Statement) Comparison period start.",
    ),
    compare_period_end: Optional[date] = Query(
        None,
        description="(Income Statement) Comparison period end.",
    ),
    # ── Shared optional ────────────────────────────────────────────────
    include_voided: bool = Query(
        False,
        description="Include voided JEs in calculations.",
    ),
    cost_center_id: Optional[List[str]] = Query(
        None,
        description=(
            "Optional — filter JE lines to one or more cost centres. "
            "Repeat the parameter for multiple values: "
            "?cost_center_id=A&cost_center_id=B."
        ),
    ),
    db: AsyncSession = Depends(get_db),
    _current_user: TokenPayload = Depends(require_roles(*_READ_ROLES)),
) -> StreamingResponse:
    """
    Export a financial statement as PDF or Excel (streaming download).

    Validation:
    - statement must be one of: balance-sheet, income-statement, cash-flow.
    - format must be: pdf or xlsx.
    - period_start / period_end required for income-statement and cash-flow.
    - Delegates data computation to the same internal functions as the
      JSON report endpoints.

    Args:
        statement: Statement slug (path param).
        format: Output format query param.
        organization_id: Org scope.
        company_code: Company code.
        as_of_date: Balance Sheet snapshot date.
        period_start: IS / CF period start.
        period_end: IS / CF period end.
        compare_period_start: IS comparison period start.
        compare_period_end: IS comparison period end.
        include_voided: Include voided JEs.
        cost_center_id: Optional list of cost-centre IDs to filter.
        db: Async DB session.
        _current_user: Authenticated user.

    Returns:
        StreamingResponse with the generated file.

    Raises:
        HTTPException 400: Invalid statement or format, or missing required params.
        HTTPException 403: Insufficient role (handled by require_roles).
        HTTPException 404: Company not found (propagated from report functions).
    """
    # ── Validate path + format params ─────────────────────────────────
    if statement not in _VALID_STATEMENTS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Invalid statement '{statement}'. "
                f"Must be one of: {', '.join(sorted(_VALID_STATEMENTS))}."
            ),
        )

    fmt = format.lower().strip()
    if fmt not in _VALID_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Invalid format '{format}'. "
                f"Must be one of: {', '.join(sorted(_VALID_FORMATS))}."
            ),
        )

    # ── Validate period params for IS / CF ────────────────────────────
    if statement in ("income-statement", "cash-flow"):
        if period_start is None or period_end is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"period_start and period_end are required for "
                    f"statement '{statement}'."
                ),
            )

    # ── Compute report data by calling the JSON endpoints' logic ──────
    # Reason: We delegate to the exact same endpoint functions so the
    # exported file is guaranteed to contain the same data as the JSON
    # view. The endpoint functions return success(ResponseModel) — we
    # unwrap the .data field from the SuccessResponse wrapper.

    if statement == "balance-sheet":
        resp = await get_balance_sheet(
            organization_id=organization_id,
            company_code=company_code,
            as_of_date=as_of_date,
            include_voided=include_voided,
            cost_center_id=cost_center_id,
            db=db,
            _current_user=_current_user,
        )
        report_data = resp.data.model_dump()
        period_label = report_data.get("asOfDate", date.today().isoformat())

    elif statement == "income-statement":
        resp = await get_income_statement(
            organization_id=organization_id,
            company_code=company_code,
            period_start=period_start,  # type: ignore[arg-type]
            period_end=period_end,       # type: ignore[arg-type]
            compare_period_start=compare_period_start,
            compare_period_end=compare_period_end,
            include_voided=include_voided,
            cost_center_id=cost_center_id,
            db=db,
            _current_user=_current_user,
        )
        report_data = resp.data.model_dump()
        period_label = (
            f"{period_start.isoformat()}_{period_end.isoformat()}"
        )

    else:  # cash-flow
        resp = await get_cash_flow(
            organization_id=organization_id,
            company_code=company_code,
            period_start=period_start,  # type: ignore[arg-type]
            period_end=period_end,       # type: ignore[arg-type]
            include_voided=include_voided,
            cost_center_id=cost_center_id,
            db=db,
            _current_user=_current_user,
        )
        report_data = resp.data.model_dump()
        period_label = (
            f"{period_start.isoformat()}_{period_end.isoformat()}"
        )

    # ── Generate file content ─────────────────────────────────────────
    if fmt == "xlsx":
        if statement == "balance-sheet":
            content = _build_balance_sheet_xlsx(report_data, company_code)
        elif statement == "income-statement":
            content = _build_income_statement_xlsx(report_data, company_code)
        else:
            content = _build_cash_flow_xlsx(report_data, company_code)

        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_ext = "xlsx"

    else:  # pdf
        if statement == "balance-sheet":
            content = _build_balance_sheet_pdf(report_data)
        elif statement == "income-statement":
            content = _build_income_statement_pdf(report_data)
        else:
            content = _build_cash_flow_pdf(report_data)

        media_type = "application/pdf"
        file_ext = "pdf"

    filename = _build_filename(statement, period_label, company_code, file_ext)

    logger.info(
        "[Finance/Export] statement=%s format=%s org=%s company=%s "
        "period=%s bytes=%d filename=%s",
        statement, fmt, organization_id, company_code,
        period_label, len(content), filename,
    )

    return StreamingResponse(
        content=io.BytesIO(content),
        media_type=media_type,
        headers={
            # Reason: 'attachment' forces browser download rather than
            # inline rendering — critical for binary file types.
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )
