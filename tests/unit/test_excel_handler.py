"""
Unit tests for the Excel import handler (import_crops and build_import_template).

All tests are pure — no real MongoDB connection.
The database call in import_crops is patched with AsyncMock.

Test cases:
  1. Points-only import — existing v1 behaviour unchanged.
  2. Net Yield-only import — converts to dripper points via yieldInfo.
  3. Both columns present — Net Yield wins, Points column is ignored.
  4. Net Yield > 0 but plant has yieldPerPlant = 0 — row skipped with reason.
  5. Net Yield that computes to > 10,000,000 points — clamped, warning emitted.
  6. Round-trip: template → parse — placeholder rows skipped, no crash.
  7. Old-format file (no Net Yield column) — parser still works as before.
  8. Net Yield is a non-numeric string — skip with "Net Yield is not a number".
  9. Net Yield is zero / negative — falls through to Points.
 10. Non-kg yieldUnit — informational warning appended.

Run:
    python -m pytest tests/unit/test_excel_handler.py -v
"""

from __future__ import annotations

import asyncio
import math
from io import BytesIO
from typing import Any, Dict, Optional
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MODULE_PATH = "src.modules.farm_manager.services.tools.excel_handler"


def _make_xlsx_bytes(
    rows: list[tuple],
    header: tuple = ("Crop Name", "Points"),
) -> bytes:
    """
    Build a minimal in-memory .xlsx file and return its bytes.

    Args:
        rows: Data rows as tuples matching the header column order.
        header: Column header labels.

    Returns:
        Raw .xlsx bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Crops"
    ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_plant_doc(
    plant_id: str,
    plant_name: str,
    yield_per_plant: float = 5.0,
    seeds_per_point: int = 1,
    waste_pct: float = 0.0,
    yield_unit: str = "kg",
) -> Dict[str, Any]:
    """Build a minimal plant_data_enhanced document with yieldInfo."""
    return {
        "plantDataId": plant_id,
        "plantName": plant_name,
        "yieldInfo": {
            "yieldPerPlant": yield_per_plant,
            "yieldUnit": yield_unit,
            "seedsPerPlantingPoint": seeds_per_point,
            "expectedWastePercentage": waste_pct,
        },
    }


def _mock_db(plant_doc: Optional[Dict[str, Any]]) -> MagicMock:
    """
    Build a mock Motor database whose plant_data_enhanced.find_one returns
    plant_doc (or None when plant_doc is None).

    Args:
        plant_doc: The document to return from find_one, or None for "not found".

    Returns:
        MagicMock database object.
    """
    db = MagicMock()
    db.plant_data_enhanced = MagicMock()
    db.plant_data_enhanced.find_one = AsyncMock(return_value=plant_doc)
    return db


def _run(coro):
    """Run a coroutine synchronously."""
    return asyncio.get_event_loop().run_until_complete(coro)


# ---------------------------------------------------------------------------
# 1. Points-only import — existing v1 behaviour
# ---------------------------------------------------------------------------

class TestPointsOnlyImport:
    """
    A two-column file (Crop Name + Points) must work exactly as before.
    This is the regression check for the original behaviour.
    """

    def test_single_crop_points_parsed(self):
        """One valid row: Tomato | 100 → items=[{plantDataId, plantName, points=100}]."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Tomato")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes([("Tomato", 100)])

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].plantName == "Tomato"
        assert result.items[0].points == 100
        assert len(result.skipped) == 0

    def test_duplicate_crops_summed(self):
        """Two rows for the same crop → points are summed."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Tomato")

        # find_one called twice, both times returns the same doc
        db = MagicMock()
        db.plant_data_enhanced = MagicMock()
        db.plant_data_enhanced.find_one = AsyncMock(side_effect=[plant_doc, plant_doc])

        file_bytes = _make_xlsx_bytes([("Tomato", 60), ("Tomato", 40)])

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 100

    def test_unknown_crop_skipped(self):
        """Crop not found in plant_data_enhanced → skipped with 'Unknown crop'."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        db = _mock_db(None)  # find_one returns None
        file_bytes = _make_xlsx_bytes([("Ghost Plant", 50)])

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 0
        assert len(result.skipped) == 1
        assert result.skipped[0].reason == "Unknown crop"

    def test_invalid_points_string_skipped(self):
        """Non-numeric Points value → skipped with 'Points is not a number'."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Potato")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes([("Potato", "abc")])

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 0
        assert len(result.skipped) == 1
        assert "not a number" in result.skipped[0].reason.lower()


# ---------------------------------------------------------------------------
# 2. Net Yield-only import
# ---------------------------------------------------------------------------

class TestNetYieldOnlyImport:
    """
    Three-column file where Points column is blank; only Net Yield is filled.
    The parser must convert Net Yield → points using the plant's yieldInfo.
    """

    def test_basic_net_yield_conversion(self):
        """
        Plant: yieldPerPlant=5, seedsPerPlantingPoint=1, waste=0%.
        yieldPerDripper = 5 * 1 * (1 - 0/100) = 5 kg/dripper.
        Net Yield = 500 kg → points = ceil(500 / 5) = 100.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(
            plant_id, "Potato",
            yield_per_plant=5.0, seeds_per_point=1, waste_pct=0.0,
        )
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Potato", None, 500)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 100
        assert len(result.skipped) == 0

    def test_waste_percentage_applied(self):
        """
        Plant: yieldPerPlant=10, seedsPerPlantingPoint=2, waste=20%.
        yieldPerDripper = 10 * 2 * (1 - 0.20) = 10 * 2 * 0.8 = 16 kg/dripper.
        Net Yield = 160 kg → points = ceil(160 / 16) = 10.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(
            plant_id, "Strawberry",
            yield_per_plant=10.0, seeds_per_point=2, waste_pct=20.0,
        )
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Strawberry", None, 160)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 10

    def test_ceil_applied_for_fractional_result(self):
        """
        yieldPerDripper = 3 kg/dripper, Net Yield = 10 → ceil(10/3) = ceil(3.333) = 4.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(
            plant_id, "Pepper",
            yield_per_plant=3.0, seeds_per_point=1, waste_pct=0.0,
        )
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Pepper", None, 10)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 4  # ceil(10 / 3)

    def test_lowercase_header_accepted(self):
        """'net yield' (lowercase) must be recognised as the Net Yield column."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Lettuce", yield_per_plant=2.0)
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Lettuce", None, 20)],
            header=("Crop Name", "Points", "net yield"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 10  # ceil(20 / 2)

    def test_header_without_unit_suffix_accepted(self):
        """'Net Yield' (without the kg suffix) must also be recognised."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Cucumber", yield_per_plant=4.0)
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Cucumber", None, 80)],
            header=("Crop Name", "Points", "Net Yield"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 20  # ceil(80 / 4)


# ---------------------------------------------------------------------------
# 3. Both columns present — Net Yield wins
# ---------------------------------------------------------------------------

class TestBothColumnsNetYieldWins:
    """When a row has both Points and Net Yield filled, Net Yield takes precedence."""

    def test_net_yield_overrides_points(self):
        """
        Points = 999 (irrelevant), Net Yield = 100 kg.
        yieldPerDripper = 10 kg/dripper → points = ceil(100/10) = 10.
        The value 999 must NOT appear in the result.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Tomato", yield_per_plant=10.0)
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Tomato", 999, 100)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 10  # from Net Yield, not 999


# ---------------------------------------------------------------------------
# 4. Net Yield > 0 but plant has yieldPerPlant = 0
# ---------------------------------------------------------------------------

class TestInvalidYieldRate:
    """When yieldPerDripper is <= 0, the row must be skipped with a clear reason."""

    def test_zero_yield_per_plant_skipped(self):
        """
        yieldPerPlant = 0 → yieldPerDripper = 0 → row skipped with
        reason "Plant has invalid yield rate".
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "BrokenPlant", yield_per_plant=0.0)
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("BrokenPlant", None, 500)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 0
        assert len(result.skipped) == 1
        assert result.skipped[0].reason == "Plant has invalid yield rate"

    def test_full_waste_percentage_skipped(self):
        """
        waste = 100% → yieldPerDripper = 0 → skip.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(
            plant_id, "WastedCrop",
            yield_per_plant=5.0, seeds_per_point=1, waste_pct=100.0,
        )
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("WastedCrop", None, 200)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 0
        assert len(result.skipped) == 1
        assert result.skipped[0].reason == "Plant has invalid yield rate"


# ---------------------------------------------------------------------------
# 5. Points clamped to maximum — warning emitted
# ---------------------------------------------------------------------------

class TestPointsClampedToMax:
    """
    When computed/given points exceed 10,000,000, clamp and emit a warning.
    The row must NOT be skipped — the conversion is the user's intent.
    """

    def test_net_yield_clamped_to_max(self):
        """
        Net Yield = 1e12, yieldPerDripper = 1 → computed points = 1e12.
        Expect: clamped to 10_000_000, warning in result.warnings.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(
            plant_id, "HyperCrop",
            yield_per_plant=1.0, seeds_per_point=1, waste_pct=0.0,
        )
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("HyperCrop", None, 1_000_000_000_000)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        # Row must not be skipped
        assert len(result.items) == 1
        assert result.items[0].points == 10_000_000

        # Warning must mention "clamped" or the max value
        clamp_warnings = [
            w for w in result.warnings
            if "clamp" in w.lower() or "10,000,000" in w or "10000000" in w
        ]
        assert len(clamp_warnings) >= 1, f"Expected clamp warning, got: {result.warnings}"

    def test_explicit_points_clamped_to_max(self):
        """
        Points column value = 20_000_000 → clamped to 10_000_000.
        """
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "TomatoMax")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("TomatoMax", 20_000_000)],
            header=("Crop Name", "Points"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 10_000_000
        assert any("clamp" in w.lower() for w in result.warnings)


# ---------------------------------------------------------------------------
# 6. Round-trip: template → parse
# ---------------------------------------------------------------------------

class TestTemplateParsing:
    """
    generate template → load it → parse it.
    Placeholder rows should be skipped (crop names won't exist in DB),
    but the parse must not crash.
    """

    def test_template_parses_without_crash(self):
        """Template round-trip: no crash, no items (placeholder names not in DB)."""
        from src.modules.farm_manager.services.tools.excel_handler import (
            build_import_template,
            import_crops,
        )

        template_bytes = build_import_template()

        # Verify it is a valid xlsx
        wb = load_workbook(BytesIO(template_bytes))
        assert "Crops" in wb.sheetnames
        ws = wb["Crops"]
        rows = list(ws.iter_rows(values_only=True))

        # Row 1 should be the header
        assert rows[0][0] == "Crop Name"
        assert rows[0][1] == "Points"
        assert rows[0][2] == "Net Yield (kg)"

        # Row 2: placeholder name + Points value of 100
        assert rows[1][1] == 100

        # Row 3: e.g. Potato + Net Yield value of 500
        assert rows[2][0] == "e.g. Potato"
        assert rows[2][2] == 500

        # parse() should run without error; DB returns None for all lookups
        db = _mock_db(None)

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(template_bytes)

        result = _run(_run_async())

        # No items because placeholder names are not in DB
        assert len(result.items) == 0
        # Placeholder rows skipped (2 data rows + 1 note row which is blank or skipped)
        # The note row at row 4 has no crop name, so it is silently skipped.
        assert len(result.skipped) <= 3

    def test_template_column_widths(self):
        """Column widths should be set correctly on the template."""
        from src.modules.farm_manager.services.tools.excel_handler import build_import_template

        template_bytes = build_import_template()
        wb = load_workbook(BytesIO(template_bytes))
        ws = wb["Crops"]

        assert ws.column_dimensions["A"].width == 36
        assert ws.column_dimensions["B"].width == 12
        assert ws.column_dimensions["C"].width == 18


# ---------------------------------------------------------------------------
# 7. Old-format file (no Net Yield column) — backward compatibility
# ---------------------------------------------------------------------------

class TestOldFormatBackwardCompatibility:
    """Two-column files (Crop Name + Points) must work without the third column."""

    def test_two_column_file_works(self):
        """Two-column import behaves exactly as before the feature."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Carrot")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Carrot", 75)],
            header=("Crop Name", "Points"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 75
        assert len(result.skipped) == 0
        assert len(result.warnings) == 0


# ---------------------------------------------------------------------------
# 8. Net Yield is a non-numeric string
# ---------------------------------------------------------------------------

class TestNetYieldNotANumber:
    """A non-numeric Net Yield value with no usable Points → skip."""

    def test_string_net_yield_skipped(self):
        """Net Yield = 'LOTS' (non-numeric) → skipped with 'Net Yield is not a number'."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Tomato")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Tomato", None, "LOTS")],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 0
        assert len(result.skipped) == 1
        assert result.skipped[0].reason == "Net Yield is not a number"


# ---------------------------------------------------------------------------
# 9. Net Yield is zero or negative — falls through to Points
# ---------------------------------------------------------------------------

class TestNetYieldZeroFallsThrough:
    """Zero / negative Net Yield is treated as absent; the Points column is checked."""

    def test_zero_net_yield_falls_to_points(self):
        """Net Yield = 0, Points = 50 → uses Points."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Onion")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Onion", 50, 0)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 50

    def test_negative_net_yield_falls_to_points(self):
        """Net Yield = -100, Points = 30 → uses Points."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Garlic")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Garlic", 30, -100)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 1
        assert result.items[0].points == 30

    def test_zero_net_yield_and_no_points_skipped(self):
        """Net Yield = 0, Points = None → skip with 'Row has neither Points nor Net Yield'."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(plant_id, "Basil")
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("Basil", None, 0)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        assert len(result.items) == 0
        assert len(result.skipped) == 1
        assert result.skipped[0].reason == "Row has neither Points nor Net Yield"


# ---------------------------------------------------------------------------
# 10. Non-kg yieldUnit — informational warning appended
# ---------------------------------------------------------------------------

class TestNonKgYieldUnitWarning:
    """When a plant has yieldUnit != 'kg', an informational warning is emitted."""

    def test_lbs_yield_unit_warning(self):
        """Plant has yieldUnit='lbs' → warning mentions 'lbs' in warnings list."""
        from src.modules.farm_manager.services.tools.excel_handler import import_crops

        plant_id = str(uuid4())
        plant_doc = _make_plant_doc(
            plant_id, "AmericanCrop",
            yield_per_plant=5.0, seeds_per_point=1, waste_pct=0.0,
            yield_unit="lbs",
        )
        db = _mock_db(plant_doc)
        file_bytes = _make_xlsx_bytes(
            [("AmericanCrop", None, 100)],
            header=("Crop Name", "Points", "Net Yield (kg)"),
        )

        async def _run_async():
            with patch(f"{_MODULE_PATH}.farm_db") as mock_farm_db:
                mock_farm_db.get_database.return_value = db
                return await import_crops(file_bytes)

        result = _run(_run_async())

        # Item should still be created (the warning is informational)
        assert len(result.items) == 1
        unit_warnings = [w for w in result.warnings if "lbs" in w]
        assert len(unit_warnings) >= 1, (
            f"Expected a 'lbs' unit warning in {result.warnings}"
        )


# ---------------------------------------------------------------------------
# 11. _is_net_yield_header helper — edge cases
# ---------------------------------------------------------------------------

class TestNetYieldHeaderDetection:
    """Unit tests for the _is_net_yield_header helper function."""

    def test_standard_form_accepted(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("Net Yield (kg)") is True

    def test_without_unit_accepted(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("Net Yield") is True

    def test_lowercase_accepted(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("net yield") is True

    def test_uppercase_accepted(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("NET YIELD (KG)") is True

    def test_other_unit_suffix_accepted(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("Net Yield (lbs)") is True

    def test_points_header_rejected(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("Points") is False

    def test_crop_name_header_rejected(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("Crop Name") is False

    def test_empty_string_rejected(self):
        from src.modules.farm_manager.services.tools.excel_handler import _is_net_yield_header
        assert _is_net_yield_header("") is False
