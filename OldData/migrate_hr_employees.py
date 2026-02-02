#!/usr/bin/env python3
"""
HR Employee Migration Script

Migrates employee data from Excel file to MongoDB employees collection.
Source: HR File - Emirates ID Staff.xlsx
Target: a64core_db.employees

Usage:
    python migrate_hr_employees.py [--dry-run] [--limit N]

Options:
    --dry-run   Preview what would be imported without making changes
    --limit N   Only process first N employees (for testing)
"""

import argparse
import asyncio
from datetime import datetime, date
from typing import Optional
from uuid import uuid4
import sys
import os

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from motor.motor_asyncio import AsyncIOMotorClient
except ImportError:
    print("Error: motor not installed. Run: pip install motor")
    sys.exit(1)


# Excel column indices (0-based)
COL_NO = 0
COL_ARABIC_FIRST = 1
COL_ARABIC_MIDDLE = 2
COL_ARABIC_LAST = 3
COL_GENDER = 4
COL_PRINCIPAL_NAME = 5
COL_EMPLOYEE_CODE = 6
COL_NATIONALITY = 7
COL_MARITAL_STATUS = 8
COL_ID_TYPE = 9
COL_NATIONAL_ID = 10
COL_VISA_PLACE = 11

# MongoDB connection
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "a64core_db")

# Default values for required fields not in Excel
DEFAULT_DEPARTMENT = "Farm Operations"
DEFAULT_POSITION = "Farm Worker"
DEFAULT_HIRE_DATE = date(2024, 1, 1)  # Placeholder - adjust as needed
DEFAULT_CREATED_BY = "00000000-0000-0000-0000-000000000000"  # System migration


def parse_english_name(principal_name: str) -> tuple[str, str]:
    """
    Parse Principal Name into firstName and lastName.

    Examples:
        "MOHAMMED SHAHED SAFIUL ALAM" -> ("MOHAMMED SHAHED SAFIUL", "ALAM")
        "MD SUJON MIAH" -> ("MD SUJON", "MIAH")
        "MUKSUDUR RAHMAN" -> ("MUKSUDUR", "RAHMAN")
    """
    if not principal_name:
        return ("Unknown", "Unknown")

    parts = principal_name.strip().split()
    if len(parts) == 1:
        return (parts[0].title(), "")
    elif len(parts) == 2:
        return (parts[0].title(), parts[1].title())
    else:
        # First N-1 parts are first name, last part is last name
        first_name = " ".join(parts[:-1]).title()
        last_name = parts[-1].title()
        return (first_name, last_name)


def normalize_gender(gender: str) -> Optional[str]:
    """Normalize gender value to enum format."""
    if not gender:
        return None
    gender_lower = gender.strip().lower()
    if gender_lower == "male":
        return "male"
    elif gender_lower == "female":
        return "female"
    return None


def normalize_marital_status(status: str) -> Optional[str]:
    """Normalize marital status to enum format."""
    if not status:
        return None
    status_lower = status.strip().lower()
    if status_lower == "single":
        return "single"
    elif status_lower == "married":
        return "married"
    elif status_lower == "divorced":
        return "divorced"
    elif status_lower == "widowed":
        return "widowed"
    return None


def generate_email(employee_code: str) -> str:
    """Generate email from employee code."""
    code = employee_code.lower().replace("-", "")
    return f"{code}@a64farms.ae"


def create_employee_document(row: tuple) -> Optional[dict]:
    """
    Create employee document from Excel row.

    Returns None if row is invalid (header or empty).
    """
    employee_code = row[COL_EMPLOYEE_CODE]
    principal_name = row[COL_PRINCIPAL_NAME]

    # Skip if no employee code or it's the header
    if not employee_code or employee_code == "Employee Code":
        return None

    first_name, last_name = parse_english_name(principal_name)

    # Generate unique ID
    employee_id = str(uuid4())

    now = datetime.utcnow()

    return {
        "employeeId": employee_id,
        "employeeCode": employee_code,

        # English name
        "firstName": first_name,
        "lastName": last_name,

        # Arabic name
        "arabicFirstName": row[COL_ARABIC_FIRST] if row[COL_ARABIC_FIRST] else None,
        "arabicMiddleName": row[COL_ARABIC_MIDDLE] if row[COL_ARABIC_MIDDLE] else None,
        "arabicLastName": row[COL_ARABIC_LAST] if row[COL_ARABIC_LAST] else None,

        # Contact
        "email": generate_email(employee_code),
        "phone": None,  # Not in Excel

        # Employment (defaults)
        "department": DEFAULT_DEPARTMENT,
        "position": DEFAULT_POSITION,
        "hireDate": datetime.combine(DEFAULT_HIRE_DATE, datetime.min.time()),
        "status": "active",

        # Personal details
        "gender": normalize_gender(row[COL_GENDER]),
        "nationality": row[COL_NATIONALITY] if row[COL_NATIONALITY] else None,
        "maritalStatus": normalize_marital_status(row[COL_MARITAL_STATUS]),

        # UAE identification
        "emiratesId": str(row[COL_NATIONAL_ID]) if row[COL_NATIONAL_ID] else None,
        "visaIssuancePlace": row[COL_VISA_PLACE] if row[COL_VISA_PLACE] else None,

        # Emergency contact (not in Excel)
        "emergencyContact": None,

        # Audit fields
        "createdBy": DEFAULT_CREATED_BY,
        "createdAt": now,
        "updatedAt": now,

        # Migration metadata
        "_migrationSource": "HR File - Emirates ID Staff.xlsx",
        "_migratedAt": now,
    }


def read_excel_employees(file_path: str, limit: Optional[int] = None) -> list[dict]:
    """
    Read employees from Excel file.

    Returns list of unique employee documents.
    """
    print(f"Reading Excel file: {file_path}")

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    employees = []
    seen_codes = set()

    # Start from row 4 (row 1-2 empty, row 3 is header)
    for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if limit and len(employees) >= limit:
            break

        employee_code = row[COL_EMPLOYEE_CODE]

        # Skip duplicates
        if employee_code in seen_codes:
            continue

        doc = create_employee_document(row)
        if doc:
            seen_codes.add(employee_code)
            employees.append(doc)

    wb.close()
    print(f"Found {len(employees)} unique employees")
    return employees


async def migrate_employees(employees: list[dict], dry_run: bool = False):
    """
    Migrate employees to MongoDB.
    """
    if dry_run:
        print("\n=== DRY RUN MODE - No changes will be made ===\n")

    print(f"Connecting to MongoDB: {MONGO_URI}")
    client = AsyncIOMotorClient(MONGO_URI)
    db = client[DB_NAME]
    collection = db.employees

    # Check for existing employees by code
    existing_codes = set()
    async for doc in collection.find({}, {"employeeCode": 1}):
        if doc.get("employeeCode"):
            existing_codes.add(doc["employeeCode"])

    print(f"Found {len(existing_codes)} existing employees in database")

    # Filter out duplicates
    new_employees = [e for e in employees if e["employeeCode"] not in existing_codes]
    skipped = len(employees) - len(new_employees)

    if skipped > 0:
        print(f"Skipping {skipped} employees (already exist)")

    if not new_employees:
        print("No new employees to import")
        return

    print(f"\nWill import {len(new_employees)} new employees")

    # Show sample
    print("\nSample records to import:")
    for emp in new_employees[:3]:
        print(f"  {emp['employeeCode']}: {emp['firstName']} {emp['lastName']}")
        print(f"    Arabic: {emp['arabicFirstName']} {emp['arabicLastName']}")
        print(f"    Email: {emp['email']}")
        print(f"    Nationality: {emp['nationality']}")
        print()

    if dry_run:
        print("=== DRY RUN COMPLETE - No changes made ===")
        return

    # Insert employees
    print("\nInserting employees...")
    result = await collection.insert_many(new_employees)
    print(f"Successfully imported {len(result.inserted_ids)} employees")

    # Create indexes if they don't exist
    print("\nCreating indexes...")
    await collection.create_index("employeeId", unique=True)
    await collection.create_index("employeeCode", unique=True)
    await collection.create_index("email", unique=True)
    await collection.create_index([
        ("firstName", "text"),
        ("lastName", "text"),
        ("arabicFirstName", "text"),
        ("arabicLastName", "text"),
        ("department", "text"),
    ])
    print("Indexes created")

    client.close()


def main():
    parser = argparse.ArgumentParser(description="Migrate HR employees from Excel to MongoDB")
    parser.add_argument("--dry-run", action="store_true", help="Preview without making changes")
    parser.add_argument("--limit", type=int, help="Limit number of employees to process")
    args = parser.parse_args()

    # Path to Excel file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "HR File - Emirates ID Staff.xlsx")

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)

    # Read employees
    employees = read_excel_employees(excel_path, limit=args.limit)

    if not employees:
        print("No employees found in Excel file")
        sys.exit(1)

    # Run migration
    asyncio.run(migrate_employees(employees, dry_run=args.dry_run))

    print("\nMigration complete!")


if __name__ == "__main__":
    main()
