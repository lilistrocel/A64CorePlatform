"""
stage10_revenue_reference.py — Build revenue_reference collection from Revenue Sheet and P&L 2024-2025.

Sources:
  1. OldData/7-April-2026/Revenue - Sheet.xlsx sheet "Revenue - Sheet"
     - Row 4-7: 4 farms × 2 periods (Aug24-Jul25 and Aug25+) × 2 channels (Silal, Market)
     - Row 3: column headers for Qty/AED/AvgPrice groupings
  2. OldData/7-April-2026/P& L Season 2024-2025.xlsx sheet "P & L Aug 24 - Jul 25"
     - Row 3: Total Sales
     - Rows 5-7: Other income items (3 docs, handled separately)
     - Row 10: Total Revenue
     - Rows 15-50: Expenditure Ledger categories

New collection: revenue_reference

Each doc has a stable deterministic UUID keyed on metadata.legacyRef.
Stage 11 adds FY2023 docs to the same collection.

Run: python stage10_revenue_reference.py [--dry-run] [--reset]
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from common import (
    DIVISION_ID,
    MIGRATION_TAG,
    ORGANIZATION_ID,
    deterministic_uuid,
    get_db,
    make_arg_parser,
    make_logger,
    print_summary,
    reset_migration_data,
    upsert_by_legacy_ref,
    utcnow,
)

STAGE = "stage10_revenue_reference"
RESET_COLLECTIONS = ["revenue_reference"]

ADMIN_UUID: str = "bff26b8f-5ce9-49b2-9126-86174eaea823"

DATA_DIR: Path = Path(__file__).parent.parent.parent.parent / "OldData" / "7-April-2026"
REVENUE_XLSX: Path = DATA_DIR / "Revenue - Sheet.xlsx"
PNL_XLSX: Path = DATA_DIR / "P& L Season 2024-2025.xlsx"

# Farm name → UUID (from migrated farms collection)
FARM_UUID_MAP: dict[str, str] = {
    "liwa 330a": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
    "liwa": "ad8ee850-2811-5290-9ef4-aee3b5f7062f",
    "al ain": "042ab6a6-74c2-58ca-84ec-73dd3186b0d9",
    "al wagen": "b2c23ace-a0db-549c-9872-28f29658a2f3",
    "al wagan": "b2c23ace-a0db-549c-9872-28f29658a2f3",
    "al khazana": "2b34823c-ac2d-58c8-a57d-ed2c8fbbdf81",
    "al khazna": "2b34823c-ac2d-58c8-a57d-ed2c8fbbdf81",
    "silal upgrade": "651103f1-1967-5261-9a38-373411f4fdfa",
}

# Period definitions
FY2024_PERIOD = {
    "start": "2024-08-01",
    "end": "2025-07-31",
    "label": "FY2024",
    "farmingYear": 2024,
}
FY2025_PERIOD = {
    "start": "2025-08-01",
    "end": "2026-07-31",
    "label": "FY2025",
    "farmingYear": 2025,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def resolve_farm(name: str) -> tuple[Optional[str], str]:
    """
    Resolve farm display name to (farmId, cleanName).

    Args:
        name: raw farm name from Excel

    Returns:
        (farmId UUID or None, clean farm name)
    """
    key = name.strip().lower()
    farm_id = FARM_UUID_MAP.get(key)
    # Normalise display name
    display = name.strip().title().replace("Liwa 330A", "Liwa 330A")
    return farm_id, display


def safe_float(val: Any) -> Optional[float]:
    """
    Convert cell value to float, returning None if not convertible.

    Args:
        val: raw cell value

    Returns:
        float or None
    """
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def build_ref_doc(
    ref_id: str,
    legacy_ref: str,
    source: str,
    period: dict,
    scope: str,
    farm_id: Optional[str],
    farm_name: Optional[str],
    channel: Optional[str],
    category: Optional[str],
    qty_kgs: Optional[float],
    amount_aed: float,
    avg_price: Optional[float],
    notes: Optional[str],
    source_row: Optional[int] = None,
) -> dict[str, Any]:
    """
    Build a revenue_reference collection document.

    Args:
        ref_id: deterministic UUID for this document
        legacy_ref: stable key for upsert
        source: source file identifier
        period: period dict with start/end/label/farmingYear
        scope: "farm" | "channel" | "total" | "other_income" | "expense_category"
        farm_id: farm UUID or None
        farm_name: farm display name
        channel: "silal" | "market_outside" | None
        category: category label for other_income / expense_category
        qty_kgs: quantity in kg (revenue only)
        amount_aed: monetary amount AED
        avg_price: average price AED/kg
        notes: free text notes
        source_row: Excel row number for traceability

    Returns:
        revenue_reference document dict
    """
    now = utcnow()
    return {
        "referenceId": ref_id,
        "source": source,
        "period": period,
        "scope": scope,
        "farmId": farm_id,
        "farmName": farm_name,
        "channel": channel,
        "category": category,
        "qtyKgs": round(qty_kgs, 3) if qty_kgs is not None else None,
        "amountAed": round(amount_aed, 2),
        "avgPriceAedPerKg": round(avg_price, 4) if avg_price is not None else None,
        "notes": notes,
        "createdBy": ADMIN_UUID,
        "divisionId": DIVISION_ID,
        "organizationId": ORGANIZATION_ID,
        "createdAt": now,
        "updatedAt": now,
        "metadata": {
            "migratedFrom": MIGRATION_TAG,
            "legacyRef": legacy_ref,
            "sourceRow": source_row,
        },
    }


# ---------------------------------------------------------------------------
# Revenue Sheet loader (16 docs: 4 farms × 2 periods × 2 channels)
# ---------------------------------------------------------------------------


def load_revenue_sheet_docs() -> list[dict[str, Any]]:
    """
    Parse Revenue - Sheet.xlsx and build revenue_reference docs.

    Column layout (0-indexed from column A):
      Col 0: Farm No
      Col 1: Farm Location
      Cols 2-4: FY2024 Silal (Qty-Kgs, AED, AvgPrice)
      Cols 5-7: FY2024 Market (Qty-Kgs, AED, AvgPrice)
      Cols 8-10: FY2025 Silal (Qty-Kgs, AED, AvgPrice)
      Cols 11-13: FY2025 Market (Qty-Kgs, AED, AvgPrice)

    Returns:
        list of revenue_reference docs
    """
    wb = openpyxl.load_workbook(REVENUE_XLSX, data_only=True)
    ws = wb["Revenue - Sheet"]
    docs: list[dict[str, Any]] = []

    # Farm data rows: 4 through 7 (1-indexed)
    for row_idx in range(4, 8):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        farm_name_raw = str(row[1]).strip() if row[1] else ""
        if not farm_name_raw:
            continue

        farm_id, farm_display = resolve_farm(farm_name_raw)

        # Define the 4 channel × period combos with their column offsets
        combos = [
            # (period_def, channel, qty_col, aed_col, avg_col)
            (FY2024_PERIOD, "silal",          2,  3,  4),
            (FY2024_PERIOD, "market_outside",  5,  6,  7),
            (FY2025_PERIOD, "silal",           8,  9, 10),
            (FY2025_PERIOD, "market_outside", 11, 12, 13),
        ]

        for period, channel, qty_col, aed_col, avg_col in combos:
            qty = safe_float(row[qty_col])
            aed = safe_float(row[aed_col])
            avg = safe_float(row[avg_col])

            if aed is None:
                aed = 0.0

            legacy_ref = (
                f"revsheet:{farm_name_raw.lower()}:"
                f"{period['label']}:{channel}"
            )
            ref_id = deterministic_uuid("revenue_ref", legacy_ref)

            doc = build_ref_doc(
                ref_id=ref_id,
                legacy_ref=legacy_ref,
                source="revenue_sheet_xlsx",
                period=period,
                scope="farm",
                farm_id=farm_id,
                farm_name=farm_display,
                channel=channel,
                category=None,
                qty_kgs=qty,
                amount_aed=aed,
                avg_price=avg,
                notes=None,
                source_row=row_idx,
            )
            docs.append(doc)

    return docs


# ---------------------------------------------------------------------------
# P&L loader
# ---------------------------------------------------------------------------


def load_pnl_docs() -> list[dict[str, Any]]:
    """
    Parse P&L 2024-2025 xlsx and build revenue_reference docs.

    Produces:
      - 1 "total" doc for Total Sales
      - 1 "total" doc for Total Revenue
      - (Other income handled by stage12_other_income.py — not included here)
      - N "expense_category" docs from Expenditure Ledger

    Returns:
        list of revenue_reference docs
    """
    wb = openpyxl.load_workbook(PNL_XLSX, data_only=True)
    ws = wb["P & L Aug 24 - Jul 25"]
    docs: list[dict[str, Any]] = []

    # Read all rows into a dict by row index for safe lookup
    row_data: dict[int, tuple] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_data[i] = row

    # ---- Total Sales (row 3, col B=1, C=2)
    total_sales_row = row_data.get(3, ())
    total_sales_val = safe_float(total_sales_row[2]) if len(total_sales_row) > 2 else None
    if total_sales_val is not None:
        legacy_ref = "pnl:fy2024:total_sales"
        doc = build_ref_doc(
            ref_id=deterministic_uuid("revenue_ref", legacy_ref),
            legacy_ref=legacy_ref,
            source="pnl_xlsx_2024_2025",
            period=FY2024_PERIOD,
            scope="total",
            farm_id=None,
            farm_name=None,
            channel=None,
            category="Total Sales",
            qty_kgs=None,
            amount_aed=total_sales_val,
            avg_price=None,
            notes="Total sales across all farms and channels",
            source_row=3,
        )
        docs.append(doc)

    # ---- Total Revenue (row 10)
    total_rev_row = row_data.get(10, ())
    total_rev_val = safe_float(total_rev_row[2]) if len(total_rev_row) > 2 else None
    if total_rev_val is not None:
        legacy_ref = "pnl:fy2024:total_revenue"
        doc = build_ref_doc(
            ref_id=deterministic_uuid("revenue_ref", legacy_ref),
            legacy_ref=legacy_ref,
            source="pnl_xlsx_2024_2025",
            period=FY2024_PERIOD,
            scope="total",
            farm_id=None,
            farm_name=None,
            channel=None,
            category="Total Revenue",
            qty_kgs=None,
            amount_aed=total_rev_val,
            avg_price=None,
            notes="Total Revenue = Total Sales + Other Income",
            source_row=10,
        )
        docs.append(doc)

    # ---- Total Expense (row 51)
    total_exp_row = row_data.get(51, ())
    total_exp_val = safe_float(total_exp_row[2]) if len(total_exp_row) > 2 else None
    if total_exp_val is not None:
        legacy_ref = "pnl:fy2024:total_expense"
        doc = build_ref_doc(
            ref_id=deterministic_uuid("revenue_ref", legacy_ref),
            legacy_ref=legacy_ref,
            source="pnl_xlsx_2024_2025",
            period=FY2024_PERIOD,
            scope="total",
            farm_id=None,
            farm_name=None,
            channel=None,
            category="Total Expense",
            qty_kgs=None,
            amount_aed=total_exp_val,
            avg_price=None,
            notes="Total expenditure for FY2024",
            source_row=51,
        )
        docs.append(doc)

    # ---- Expenditure Ledger (rows 15-50)
    for row_idx in range(15, 52):
        row = row_data.get(row_idx, ())
        if len(row) < 3:
            continue
        period_col = row[0]
        particulars = str(row[1]).strip() if row[1] else ""
        amount = safe_float(row[2])

        if not particulars or amount is None:
            continue
        if particulars.lower() in ("particulars", "total expense", "expenditure ledger"):
            continue
        if particulars.startswith("Period"):
            continue

        # Reason: Skip total/summary rows that are already captured above
        if particulars.lower().startswith("total"):
            continue

        legacy_ref = f"pnl:fy2024:expense:{particulars.lower().replace(' ', '_')}"
        doc = build_ref_doc(
            ref_id=deterministic_uuid("revenue_ref", legacy_ref),
            legacy_ref=legacy_ref,
            source="pnl_xlsx_2024_2025",
            period=FY2024_PERIOD,
            scope="expense_category",
            farm_id=None,
            farm_name=None,
            channel=None,
            category=particulars,
            qty_kgs=None,
            amount_aed=amount,
            avg_price=None,
            notes=f"Period: {period_col}" if period_col and str(period_col).strip().lower() != "year" else None,
            source_row=row_idx,
        )
        docs.append(doc)

    return docs


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def run(dry_run: bool, reset: bool) -> None:
    """
    Main entry point for stage 10.

    Args:
        dry_run: if True, read and compute but do not write
        reset: if True, delete all migration-tagged revenue_reference docs
    """
    logger = make_logger(STAGE)
    db = get_db()

    if reset:
        logger.info("[RESET] Deleting migration-tagged revenue_reference docs...")
        reset_migration_data(db, RESET_COLLECTIONS, logger)

    if not dry_run:
        db.revenue_reference.create_index("metadata.legacyRef", unique=False, background=True)

    # -----------------------------------------------------------------------
    # Load docs from both sources
    # -----------------------------------------------------------------------
    logger.info("Loading Revenue Sheet (4 farms × 2 periods × 2 channels)...")
    rev_docs = load_revenue_sheet_docs()
    logger.info(f"  Revenue sheet docs: {len(rev_docs)} (expected 16)")

    logger.info("Loading P&L 2024-2025 (totals + expenditure categories)...")
    pnl_docs = load_pnl_docs()
    logger.info(f"  P&L docs: {len(pnl_docs)} (expected ~40)")

    all_docs = rev_docs + pnl_docs
    logger.info(f"Total docs to upsert: {len(all_docs)}")

    # -----------------------------------------------------------------------
    # Dry-run preview
    # -----------------------------------------------------------------------
    if dry_run:
        logger.info("[DRY-RUN] Preview of docs that would be inserted:")
        for doc in all_docs:
            logger.info(
                f"  scope={doc['scope']} farm={doc.get('farmName')} "
                f"channel={doc.get('channel')} cat={doc.get('category')} "
                f"period={doc['period']['label']} AED={doc['amountAed']:,.2f}"
            )

    # -----------------------------------------------------------------------
    # Upsert
    # -----------------------------------------------------------------------
    inserted = updated = 0
    error_samples: list[str] = []

    for doc in all_docs:
        legacy_ref = doc["metadata"]["legacyRef"]
        try:
            ins, upd = upsert_by_legacy_ref(
                db.revenue_reference, doc, legacy_ref, dry_run=dry_run, logger=logger
            )
            if ins:
                inserted += 1
            elif upd:
                updated += 1
        except Exception as exc:
            msg = f"Upsert failed [{legacy_ref}]: {exc}"
            logger.error(msg)
            error_samples.append(msg)

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    logger.info("")
    logger.info("=" * 60)
    logger.info("  REVENUE REFERENCE SUMMARY (FY2024):")
    rev_farm_docs = [d for d in rev_docs if d["scope"] == "farm"]
    logger.info(f"    Farm × channel revenue docs: {len(rev_farm_docs)}")
    total_rev_aed = sum(
        d["amountAed"] for d in rev_farm_docs if d.get("channel") and d["amountAed"] > 0
    )
    logger.info(f"    Total revenue in revenue sheet: AED {total_rev_aed:,.2f}")
    logger.info(f"    P&L docs (totals + expense cats): {len(pnl_docs)}")
    exp_docs = [d for d in pnl_docs if d["scope"] == "expense_category"]
    total_exp = sum(d["amountAed"] for d in exp_docs)
    logger.info(f"    Total expense categories: {len(exp_docs)} = AED {total_exp:,.2f}")
    logger.info("=" * 60)

    print_summary(
        stage=STAGE,
        rows_read=len(all_docs),
        rows_inserted=inserted,
        rows_updated=updated,
        rows_skipped=0,
        rows_errored=len(error_samples),
        error_samples=error_samples,
        logger=logger,
    )


if __name__ == "__main__":
    parser = make_arg_parser(
        "Stage 10: Build revenue_reference collection from Revenue Sheet and P&L 2024-2025"
    )
    args = parser.parse_args()
    run(dry_run=args.dry_run, reset=args.reset)
